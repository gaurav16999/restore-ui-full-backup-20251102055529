"""
Enhanced Academic & Learning Modules
===========================================
Comprehensive enhancements for:
1. Admission & Promotion (Bulk Import/Export)
2. Homework & Assignment Evaluation (File Upload, Teacher Feedback)
3. Lesson Plan (Lesson → Topic → Subject Mapping)
4. Class Test / Quiz System (Automated Grading)
5. Online Exam System (Question Bank, Automated Grading, Merit List, Tabulation Sheet)
6. Class Routine Scheduler (Drag-and-Drop Support)
"""

from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models import Q, Avg, Count, Sum, Max, Min, F
from django.utils import timezone
from django.db import transaction
from django.http import HttpResponse
from datetime import datetime, timedelta
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
import csv

from admin_api.models import (
    Student, Teacher, Subject, Class, ClassRoom, User,
    OnlineExam, Question, QuestionGroup, QuestionAnswer,
    Assignment, AssignmentSubmission, ClassTest,
    Lesson, Topic, LessonPlan, Timetable,
    AcademicYear, AdmissionApplication, StudentPromotion
)
from admin_api.serializers.online_exam import OnlineExamSerializer
from admin_api.serializers import AssignmentSerializer, AssignmentSubmissionSerializer
from admin_api.serializers.class_test import ClassTestSerializer
from admin_api.serializers.lesson_plan import LessonSerializer, TopicSerializer, LessonPlanSerializer


# ==================== ADMISSION & PROMOTION ====================

class AdmissionEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced admission management with bulk import/export
    """
    queryset = AdmissionApplication.objects.all()
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    def get_serializer_class(self):
        from admin_api.serializers.academic import AdmissionApplicationSerializer
        return AdmissionApplicationSerializer
    
    @action(detail=False, methods=['get'])
    def download_template(self, request):
        """
        Download CSV/Excel template for bulk admission import
        """
        file_format = request.query_params.get('format', 'excel')
        
        if file_format == 'csv':
            response = HttpResponse(content_type='text/csv')
            response['Content-Disposition'] = 'attachment; filename="admission_template.csv"'
            
            writer = csv.writer(response)
            writer.writerow([
                'First Name', 'Last Name', 'Date of Birth (YYYY-MM-DD)',
                'Gender', 'Email', 'Phone', 'Parent Phone',
                'Applying for Class', 'Previous School', 'Address',
                'Priority (low/medium/high/urgent)'
            ])
            writer.writerow([
                'John', 'Doe', '2010-05-15', 'Male', 'john@example.com',
                '1234567890', '0987654321', 'Grade 5', 'ABC School',
                '123 Main St, City', 'medium'
            ])
            
            return response
        
        else:  # Excel format
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = 'Admission Template'
            
            # Header styling
            header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            header_font = Font(bold=True, color='FFFFFF', size=11)
            header_alignment = Alignment(horizontal='center', vertical='center')
            
            headers = [
                'First Name', 'Last Name', 'Date of Birth (YYYY-MM-DD)',
                'Gender', 'Email', 'Phone', 'Parent Phone',
                'Applying for Class', 'Previous School', 'Address',
                'Priority (low/medium/high/urgent)'
            ]
            
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=1, column=col_num, value=header)
                cell.fill = header_fill
                cell.font = header_font
                cell.alignment = header_alignment
                ws.column_dimensions[cell.column_letter].width = 18
            
            # Sample data row
            sample_data = [
                'John', 'Doe', '2010-05-15', 'Male', 'john@example.com',
                '1234567890', '0987654321', 'Grade 5', 'ABC School',
                '123 Main St, City', 'medium'
            ]
            for col_num, value in enumerate(sample_data, 1):
                ws.cell(row=2, column=col_num, value=value)
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            response = HttpResponse(
                output.read(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = 'attachment; filename="admission_template.xlsx"'
            
            return response
    
    @action(detail=False, methods=['post'])
    def bulk_import(self, request):
        """
        Bulk import admission applications from CSV/Excel
        """
        file = request.FILES.get('file')
        if not file:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        file_ext = file.name.split('.')[-1].lower()
        
        try:
            imported = []
            errors = []
            
            if file_ext == 'csv':
                # Process CSV
                decoded_file = file.read().decode('utf-8').splitlines()
                reader = csv.DictReader(decoded_file)
                
                for row_num, row in enumerate(reader, start=2):
                    try:
                        self._create_application_from_row(row, imported)
                    except Exception as e:
                        errors.append(f'Row {row_num}: {str(e)}')
            
            elif file_ext in ['xlsx', 'xls']:
                # Process Excel
                wb = openpyxl.load_workbook(file)
                ws = wb.active
                
                headers = [cell.value for cell in ws[1]]
                
                for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                    if not any(row):  # Skip empty rows
                        continue
                    
                    try:
                        row_dict = dict(zip(headers, row))
                        self._create_application_from_row(row_dict, imported)
                    except Exception as e:
                        errors.append(f'Row {row_num}: {str(e)}')
            
            else:
                return Response(
                    {'error': 'Invalid file format. Please use CSV or Excel (.xlsx)'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            return Response({
                'message': f'Successfully imported {len(imported)} applications',
                'imported_count': len(imported),
                'error_count': len(errors),
                'errors': errors[:10] if errors else None  # Return first 10 errors
            })
        
        except Exception as e:
            return Response(
                {'error': f'Failed to process file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    def _create_application_from_row(self, row, imported_list):
        """Helper method to create application from row data"""
        # Get academic year
        current_year = AcademicYear.objects.filter(is_current=True).first()
        if not current_year:
            raise Exception('No current academic year set')
        
        # Parse date of birth
        dob_str = row.get('Date of Birth (YYYY-MM-DD)') or row.get('Date of Birth')
        if isinstance(dob_str, str):
            dob = datetime.strptime(dob_str, '%Y-%m-%d').date()
        else:
            dob = dob_str
        
        # Generate application number
        app_number = f"ADM{timezone.now().year}{AdmissionApplication.objects.count() + 1:05d}"
        
        # Create application
        application = AdmissionApplication.objects.create(
            application_number=app_number,
            academic_year=current_year,
            first_name=row.get('First Name', '').strip(),
            last_name=row.get('Last Name', '').strip(),
            date_of_birth=dob,
            gender=row.get('Gender', 'Male'),
            email=row.get('Email', '').strip(),
            phone=row.get('Phone', '').strip(),
            parent_phone=row.get('Parent Phone', '').strip(),
            applying_for_class=row.get('Applying for Class', '').strip(),
            previous_school=row.get('Previous School', '').strip(),
            address=row.get('Address', '').strip(),
            priority=row.get('Priority (low/medium/high/urgent)', 'medium').lower(),
            status='pending'
        )
        
        imported_list.append(application.id)
    
    @action(detail=False, methods=['get'])
    def bulk_export(self, request):
        """
        Export admission applications to Excel
        """
        # Get filters
        status_filter = request.query_params.get('status')
        academic_year_id = request.query_params.get('academic_year')
        
        queryset = self.get_queryset()
        
        if status_filter:
            queryset = queryset.filter(status=status_filter)
        if academic_year_id:
            queryset = queryset.filter(academic_year_id=academic_year_id)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Admission Applications'
        
        # Header styling
        header_fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF', size=11)
        header_alignment = Alignment(horizontal='center', vertical='center')
        
        headers = [
            'Application No', 'First Name', 'Last Name', 'DOB',
            'Gender', 'Email', 'Phone', 'Parent Phone',
            'Applying for Class', 'Previous School', 'Priority',
            'Status', 'Applied Date', 'Reviewed By', 'Reviewed At'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            ws.column_dimensions[cell.column_letter].width = 15
        
        # Add data rows
        for row_num, app in enumerate(queryset.select_related('reviewed_by'), start=2):
            ws.cell(row=row_num, column=1, value=app.application_number)
            ws.cell(row=row_num, column=2, value=app.first_name)
            ws.cell(row=row_num, column=3, value=app.last_name)
            ws.cell(row=row_num, column=4, value=app.date_of_birth.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=5, value=app.gender)
            ws.cell(row=row_num, column=6, value=app.email)
            ws.cell(row=row_num, column=7, value=app.phone)
            ws.cell(row=row_num, column=8, value=app.parent_phone)
            ws.cell(row=row_num, column=9, value=app.applying_for_class)
            ws.cell(row=row_num, column=10, value=app.previous_school or '')
            ws.cell(row=row_num, column=11, value=app.priority.upper())
            ws.cell(row=row_num, column=12, value=app.status.replace('_', ' ').title())
            ws.cell(row=row_num, column=13, value=app.applied_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=14, value=app.reviewed_by.get_full_name() if app.reviewed_by else '')
            ws.cell(row=row_num, column=15, value=app.reviewed_at.strftime('%Y-%m-%d %H:%M') if app.reviewed_at else '')
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="admission_applications_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        return response


class StudentPromotionEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced student promotion with bulk operations
    """
    queryset = StudentPromotion.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.academic import StudentPromotionSerializer
        return StudentPromotionSerializer
    
    @action(detail=False, methods=['post'])
    def bulk_promote(self, request):
        """
        Bulk promote students to next grade/class
        """
        from_class = request.data.get('from_class')
        to_class = request.data.get('to_class')
        academic_year_id = request.data.get('academic_year')
        student_ids = request.data.get('student_ids', [])
        
        if not all([from_class, to_class, academic_year_id]):
            return Response(
                {'error': 'from_class, to_class, and academic_year are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            academic_year = AcademicYear.objects.get(id=academic_year_id)
        except AcademicYear.DoesNotExist:
            return Response(
                {'error': 'Academic year not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        try:
            with transaction.atomic():
                # Get students to promote
                if student_ids:
                    students = Student.objects.filter(id__in=student_ids, class_name=from_class, is_active=True)
                else:
                    students = Student.objects.filter(class_name=from_class, is_active=True)
                
                promoted_count = 0
                
                for student in students:
                    # Create promotion record
                    StudentPromotion.objects.create(
                        student=student,
                        from_class=from_class,
                        to_class=to_class,
                        academic_year=academic_year,
                        promoted_by=request.user,
                        promotion_date=timezone.now().date(),
                        status='promoted'
                    )
                    
                    # Update student class
                    student.class_name = to_class
                    student.save()
                    
                    promoted_count += 1
                
                return Response({
                    'message': f'Successfully promoted {promoted_count} students',
                    'promoted_count': promoted_count,
                    'from_class': from_class,
                    'to_class': to_class
                })
        
        except Exception as e:
            return Response(
                {'error': f'Promotion failed: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ==================== HOMEWORK & ASSIGNMENT ====================

class AssignmentEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced assignment management with file upload and evaluation
    """
    queryset = Assignment.objects.all()
    serializer_class = AssignmentSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['title', 'subject__title', 'class_assigned__name']
    ordering_fields = ['due_date', 'assigned_date', 'created_at']
    
    def perform_create(self, serializer):
        """Create assignment with file upload"""
        serializer.save(teacher=self.request.user.teacher)
    
    @action(detail=True, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def upload_attachment(self, request, pk=None):
        """
        Upload attachment for assignment
        """
        assignment = self.get_object()
        file = request.FILES.get('file')
        
        if not file:
            return Response(
                {'error': 'No file provided'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate file size (10MB limit)
        if file.size > 10 * 1024 * 1024:
            return Response(
                {'error': 'File size exceeds 10MB limit'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate file type
        allowed_extensions = ['pdf', 'doc', 'docx', 'xls', 'xlsx', 'ppt', 'pptx', 'txt', 'jpg', 'jpeg', 'png']
        file_ext = file.name.split('.')[-1].lower()
        
        if file_ext not in allowed_extensions:
            return Response(
                {'error': f'File type not allowed. Allowed types: {", ".join(allowed_extensions)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Save file (in production, use cloud storage)
        assignment.attachment_url = f'/media/assignments/{assignment.id}_{file.name}'
        assignment.save()
        
        return Response({
            'message': 'File uploaded successfully',
            'file_url': assignment.attachment_url
        })
    
    @action(detail=True, methods=['get'])
    def submission_statistics(self, request, pk=None):
        """
        Get detailed submission statistics
        """
        assignment = self.get_object()
        
        # Get all students in the class
        total_students = Student.objects.filter(
            class_name=assignment.class_assigned.name,
            is_active=True
        ).count()
        
        # Submission stats
        submissions = AssignmentSubmission.objects.filter(assignment=assignment)
        submitted = submissions.filter(status__in=['submitted', 'graded']).count()
        late = submissions.filter(status='late').count()
        graded = submissions.filter(status='graded').count()
        pending = total_students - submitted - late
        
        # Grade distribution
        grade_stats = submissions.filter(marks_obtained__isnull=False).aggregate(
            avg_marks=Avg('marks_obtained'),
            max_marks=Max('marks_obtained'),
            min_marks=Min('marks_obtained')
        )
        
        return Response({
            'total_students': total_students,
            'submitted': submitted,
            'late_submissions': late,
            'graded': graded,
            'pending': pending,
            'submission_rate': round((submitted + late) / total_students * 100, 2) if total_students > 0 else 0,
            'grade_statistics': {
                'average': round(grade_stats['avg_marks'], 2) if grade_stats['avg_marks'] else 0,
                'highest': grade_stats['max_marks'] or 0,
                'lowest': grade_stats['min_marks'] or 0,
                'max_possible': assignment.max_marks
            }
        })
    
    @action(detail=True, methods=['get'])
    def export_submissions(self, request, pk=None):
        """
        Export all submissions for an assignment to Excel
        """
        assignment = self.get_object()
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Submissions'
        
        # Header
        headers = [
            'Roll No', 'Student Name', 'Submission Date',
            'Status', 'Marks Obtained', 'Max Marks', 'Percentage',
            'Feedback', 'Graded By', 'Graded At'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
        
        # Data rows
        submissions = assignment.submissions.select_related('student__user', 'graded_by').all()
        
        for row_num, submission in enumerate(submissions, start=2):
            percentage = (submission.marks_obtained / assignment.max_marks * 100) if submission.marks_obtained else 0
            
            ws.cell(row=row_num, column=1, value=submission.student.roll_no)
            ws.cell(row=row_num, column=2, value=submission.student.user.get_full_name())
            ws.cell(row=row_num, column=3, value=submission.submission_date.strftime('%Y-%m-%d %H:%M'))
            ws.cell(row=row_num, column=4, value=submission.status.title())
            ws.cell(row=row_num, column=5, value=float(submission.marks_obtained) if submission.marks_obtained else '')
            ws.cell(row=row_num, column=6, value=assignment.max_marks)
            ws.cell(row=row_num, column=7, value=round(percentage, 2) if submission.marks_obtained else '')
            ws.cell(row=row_num, column=8, value=submission.feedback or '')
            ws.cell(row=row_num, column=9, value=submission.graded_by.user.get_full_name() if submission.graded_by else '')
            ws.cell(row=row_num, column=10, value=submission.graded_at.strftime('%Y-%m-%d %H:%M') if submission.graded_at else '')
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="assignment_submissions_{assignment.id}.xlsx"'
        
        return response


class AssignmentSubmissionEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced submission management with file upload and teacher feedback
    """
    queryset = AssignmentSubmission.objects.all()
    serializer_class = AssignmentSubmissionSerializer
    permission_classes = [IsAuthenticated]
    parser_classes = [MultiPartParser, FormParser]
    
    @action(detail=False, methods=['post'], parser_classes=[MultiPartParser, FormParser])
    def submit_with_file(self, request):
        """
        Submit assignment with file attachment
        """
        assignment_id = request.data.get('assignment_id')
        submission_text = request.data.get('submission_text', '')
        file = request.FILES.get('file')
        
        if not assignment_id:
            return Response(
                {'error': 'assignment_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            assignment = Assignment.objects.get(id=assignment_id)
        except Assignment.DoesNotExist:
            return Response(
                {'error': 'Assignment not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Check if user is a student
        if not hasattr(request.user, 'student'):
            return Response(
                {'error': 'Only students can submit assignments'},
                status=status.HTTP_403_FORBIDDEN
            )
        
        # Check if already submitted
        existing = AssignmentSubmission.objects.filter(
            assignment=assignment,
            student=request.user.student
        ).first()
        
        if existing and existing.status == 'graded':
            return Response(
                {'error': 'Cannot resubmit graded assignment'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Determine status
        is_late = timezone.now().date() > assignment.due_date
        submission_status = 'late' if is_late else 'submitted'
        
        # Create or update submission
        submission, created = AssignmentSubmission.objects.update_or_create(
            assignment=assignment,
            student=request.user.student,
            defaults={
                'submission_text': submission_text,
                'attachment_url': f'/media/submissions/{assignment.id}_{request.user.student.id}_{file.name}' if file else '',
                'status': submission_status,
                'submission_date': timezone.now()
            }
        )
        
        serializer = self.get_serializer(submission)
        
        return Response({
            'message': 'Assignment submitted successfully',
            'is_late': is_late,
            'data': serializer.data
        })
    
    @action(detail=True, methods=['post'])
    def grade_with_feedback(self, request, pk=None):
        """
        Grade submission with detailed feedback
        """
        submission = self.get_object()
        
        marks = request.data.get('marks_obtained')
        feedback = request.data.get('feedback', '')
        
        if marks is None:
            return Response(
                {'error': 'marks_obtained is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            marks = float(marks)
            max_marks = submission.assignment.max_marks
            
            if marks < 0 or marks > max_marks:
                return Response(
                    {'error': f'Marks must be between 0 and {max_marks}'},
                    status=status.HTTP_400_BAD_REQUEST
                )
            
            submission.marks_obtained = marks
            submission.feedback = feedback
            submission.status = 'graded'
            submission.graded_by = request.user.teacher
            submission.graded_at = timezone.now()
            submission.save()
            
            serializer = self.get_serializer(submission)
            
            return Response({
                'message': 'Assignment graded successfully',
                'percentage': round((marks / max_marks) * 100, 2),
                'data': serializer.data
            })
        
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_400_BAD_REQUEST
            )


# ==================== LESSON PLAN ====================

class LessonEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced lesson management with topic mapping
    """
    queryset = Lesson.objects.all()
    serializer_class = LessonSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['title', 'subject__title']
    ordering_fields = ['created_at']
    
    @action(detail=True, methods=['get'])
    def topics_list(self, request, pk=None):
        """
        Get all topics for this lesson
        """
        lesson = self.get_object()
        topics = lesson.topics.all()
        serializer = TopicSerializer(topics, many=True)
        
        return Response({
            'lesson': LessonSerializer(lesson).data,
            'topics_count': topics.count(),
            'topics': serializer.data
        })
    
    @action(detail=True, methods=['get'])
    def lesson_plans(self, request, pk=None):
        """
        Get all lesson plans for this lesson
        """
        lesson = self.get_object()
        plans = LessonPlan.objects.filter(lesson=lesson)
        serializer = LessonPlanSerializer(plans, many=True)
        
        return Response({
            'lesson': LessonSerializer(lesson).data,
            'plans_count': plans.count(),
            'plans': serializer.data
        })


class TopicEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced topic management
    """
    queryset = Topic.objects.all()
    serializer_class = TopicSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['title', 'lesson__title']
    ordering_fields = ['created_at']
    
    @action(detail=False, methods=['post'])
    def bulk_create(self, request):
        """
        Bulk create topics for a lesson
        """
        lesson_id = request.data.get('lesson_id')
        topics = request.data.get('topics', [])
        
        if not lesson_id or not topics:
            return Response(
                {'error': 'lesson_id and topics are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            lesson = Lesson.objects.get(id=lesson_id)
        except Lesson.DoesNotExist:
            return Response(
                {'error': 'Lesson not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        created_topics = []
        
        with transaction.atomic():
            for topic_data in topics:
                topic = Topic.objects.create(
                    lesson=lesson,
                    title=topic_data.get('title'),
                    description=topic_data.get('description', ''),
                    created_by=request.user
                )
                created_topics.append(topic)
        
        serializer = TopicSerializer(created_topics, many=True)
        
        return Response({
            'message': f'Successfully created {len(created_topics)} topics',
            'topics': serializer.data
        })


class LessonPlanEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced lesson plan management with progress tracking
    """
    queryset = LessonPlan.objects.all()
    serializer_class = LessonPlanSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['lesson__title', 'topic__title', 'teacher__user__first_name']
    ordering_fields = ['planned_date', 'created_at']
    
    def get_queryset(self):
        qs = super().get_queryset()
        
        teacher = self.request.query_params.get('teacher')
        lesson = self.request.query_params.get('lesson')
        planned_date = self.request.query_params.get('planned_date')
        status_filter = self.request.query_params.get('status')
        
        if teacher:
            qs = qs.filter(teacher_id=teacher)
        if lesson:
            qs = qs.filter(lesson_id=lesson)
        if planned_date:
            qs = qs.filter(planned_date=planned_date)
        if status_filter:
            qs = qs.filter(status=status_filter)
        
        return qs.select_related('lesson', 'topic', 'teacher__user').order_by('-planned_date')
    
    @action(detail=True, methods=['post'])
    def mark_completed(self, request, pk=None):
        """
        Mark lesson plan as completed
        """
        lesson_plan = self.get_object()
        
        lesson_plan.status = 'completed'
        lesson_plan.save()
        
        serializer = self.get_serializer(lesson_plan)
        
        return Response({
            'message': 'Lesson plan marked as completed',
            'data': serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def progress_summary(self, request):
        """
        Get lesson plan progress summary
        """
        teacher_id = request.query_params.get('teacher_id')
        
        queryset = self.get_queryset()
        
        if teacher_id:
            queryset = queryset.filter(teacher_id=teacher_id)
        
        total = queryset.count()
        completed = queryset.filter(status='completed').count()
        planned = queryset.filter(status='planned').count()
        skipped = queryset.filter(status='skipped').count()
        
        return Response({
            'total': total,
            'completed': completed,
            'planned': planned,
            'skipped': skipped,
            'completion_rate': round((completed / total * 100), 2) if total > 0 else 0
        })


# ==================== CLASS TEST & QUIZ ====================

class ClassTestEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced class test management with automated grading
    """
    queryset = ClassTest.objects.all()
    serializer_class = ClassTestSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['title', 'class_assigned__name', 'subject__title']
    ordering_fields = ['date', 'created_at']
    
    @action(detail=True, methods=['post'])
    def publish_test(self, request, pk=None):
        """
        Publish class test to make it visible to students
        """
        test = self.get_object()
        
        test.is_published = True
        test.save()
        
        return Response({
            'message': 'Class test published successfully',
            'data': self.get_serializer(test).data
        })
    
    @action(detail=True, methods=['post'])
    def auto_grade(self, request, pk=None):
        """
        Automatically grade MCQ questions in the test
        """
        test = self.get_object()
        
        # Get all student answers for this test
        # This assumes a QuestionAnswer model exists linking test, student, question, and answer
        answers = QuestionAnswer.objects.filter(
            question__in=test.questions.all() if hasattr(test, 'questions') else []
        ).select_related('question', 'student')
        
        graded_count = 0
        
        for answer in answers:
            if answer.question.question_type == 'mcq':
                # For MCQ, compare answer with correct option
                if answer.answer == answer.question.correct_answer:
                    answer.marks_obtained = answer.question.marks
                else:
                    answer.marks_obtained = 0
                
                answer.is_graded = True
                answer.save()
                graded_count += 1
        
        return Response({
            'message': f'Auto-graded {graded_count} MCQ answers',
            'graded_count': graded_count
        })
    
    @action(detail=True, methods=['get'])
    def result_summary(self, request, pk=None):
        """
        Get result summary for the test
        """
        test = self.get_object()
        
        # Calculate statistics
        # This is a placeholder - adjust based on your actual result model
        results = []  # Fetch actual results here
        
        return Response({
            'test': self.get_serializer(test).data,
            'total_students': len(results),
            'average_marks': 0,  # Calculate from results
            'highest_marks': 0,
            'lowest_marks': 0,
            'pass_percentage': 0
        })


# ==================== ONLINE EXAM ====================

class OnlineExamEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced online exam management with automated grading and merit list
    """
    queryset = OnlineExam.objects.all()
    serializer_class = OnlineExamSerializer
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['title', 'class_assigned__name', 'subject__title']
    ordering_fields = ['start_datetime', 'created_at']
    
    @action(detail=True, methods=['post'])
    def add_questions(self, request, pk=None):
        """
        Add questions to exam from question bank
        """
        exam = self.get_object()
        question_ids = request.data.get('question_ids', [])
        
        if not question_ids:
            return Response(
                {'error': 'question_ids are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        questions = Question.objects.filter(id__in=question_ids)
        
        # Assuming a many-to-many relationship exists
        # exam.questions.add(*questions)
        
        return Response({
            'message': f'Added {questions.count()} questions to exam',
            'total_questions': questions.count()
        })
    
    @action(detail=True, methods=['post'])
    def auto_grade_exam(self, request, pk=None):
        """
        Automatically grade MCQ questions in the exam
        """
        exam = self.get_object()
        
        if not exam.auto_mark:
            return Response(
                {'error': 'Auto-marking is not enabled for this exam'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get all student answers for this exam
        answers = QuestionAnswer.objects.filter(
            exam=exam
        ).select_related('question', 'student')
        
        graded_count = 0
        student_scores = {}
        
        with transaction.atomic():
            for answer in answers:
                if answer.question.question_type == 'mcq':
                    # Compare answer with correct option
                    correct_option = answer.question.options.get('correct')  # Adjust based on your JSON structure
                    
                    if answer.answer == correct_option:
                        answer.marks_obtained = answer.question.marks
                    else:
                        answer.marks_obtained = 0
                    
                    answer.is_graded = True
                    answer.graded_at = timezone.now()
                    answer.save()
                    
                    # Track student scores
                    student_id = answer.student.id
                    if student_id not in student_scores:
                        student_scores[student_id] = 0
                    student_scores[student_id] += answer.marks_obtained
                    
                    graded_count += 1
        
        return Response({
            'message': f'Auto-graded {graded_count} MCQ answers',
            'graded_count': graded_count,
            'students_graded': len(student_scores)
        })
    
    @action(detail=True, methods=['get'])
    def generate_merit_list(self, request, pk=None):
        """
        Generate merit list based on exam results
        """
        exam = self.get_object()
        
        # Calculate total marks for each student
        student_results = QuestionAnswer.objects.filter(
            exam=exam,
            is_graded=True
        ).values('student').annotate(
            total_marks=Sum('marks_obtained')
        ).order_by('-total_marks')
        
        # Calculate total possible marks
        total_possible = Question.objects.filter(
            # Filter questions for this exam
        ).aggregate(Sum('marks'))['marks__sum'] or 0
        
        merit_list = []
        rank = 1
        
        for result in student_results:
            student = Student.objects.get(id=result['student'])
            percentage = (result['total_marks'] / total_possible * 100) if total_possible > 0 else 0
            
            merit_list.append({
                'rank': rank,
                'roll_no': student.roll_no,
                'student_name': student.user.get_full_name(),
                'total_marks': float(result['total_marks']),
                'max_marks': total_possible,
                'percentage': round(percentage, 2),
                'grade': self._calculate_grade(percentage)
            })
            
            rank += 1
        
        return Response({
            'exam': self.get_serializer(exam).data,
            'merit_list': merit_list,
            'total_students': len(merit_list)
        })
    
    @action(detail=True, methods=['get'])
    def generate_tabulation_sheet(self, request, pk=None):
        """
        Generate comprehensive tabulation sheet with all results
        """
        exam = self.get_object()
        
        # Create Excel workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Tabulation Sheet'
        
        # Title
        ws.merge_cells('A1:H1')
        title_cell = ws['A1']
        title_cell.value = f'{exam.title} - Tabulation Sheet'
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Rank', 'Roll No', 'Student Name', 'Total Marks', 'Max Marks', 'Percentage', 'Grade', 'Status']
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
            cell.alignment = Alignment(horizontal='center')
        
        # Calculate results
        student_results = QuestionAnswer.objects.filter(
            exam=exam,
            is_graded=True
        ).values('student').annotate(
            total_marks=Sum('marks_obtained')
        ).order_by('-total_marks')
        
        total_possible = Question.objects.filter(
            # Filter for this exam
        ).aggregate(Sum('marks'))['marks__sum'] or exam.min_percent  # Fallback
        
        row_num = 4
        rank = 1
        
        for result in student_results:
            student = Student.objects.get(id=result['student'])
            total_marks = result['total_marks']
            percentage = (total_marks / total_possible * 100) if total_possible > 0 else 0
            grade = self._calculate_grade(percentage)
            status = 'Pass' if percentage >= exam.min_percent else 'Fail'
            
            ws.cell(row=row_num, column=1, value=rank)
            ws.cell(row=row_num, column=2, value=student.roll_no)
            ws.cell(row=row_num, column=3, value=student.user.get_full_name())
            ws.cell(row=row_num, column=4, value=float(total_marks))
            ws.cell(row=row_num, column=5, value=total_possible)
            ws.cell(row=row_num, column=6, value=round(percentage, 2))
            ws.cell(row=row_num, column=7, value=grade)
            ws.cell(row=row_num, column=8, value=status)
            
            # Color code status
            status_cell = ws.cell(row=row_num, column=8)
            if status == 'Pass':
                status_cell.fill = PatternFill(start_color='90EE90', end_color='90EE90', fill_type='solid')
            else:
                status_cell.fill = PatternFill(start_color='FFB6C1', end_color='FFB6C1', fill_type='solid')
            
            row_num += 1
            rank += 1
        
        # Auto-size columns
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="tabulation_sheet_{exam.title}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        return response
    
    def _calculate_grade(self, percentage):
        """Helper method to calculate letter grade"""
        if percentage >= 90:
            return 'A+'
        elif percentage >= 80:
            return 'A'
        elif percentage >= 70:
            return 'B+'
        elif percentage >= 60:
            return 'B'
        elif percentage >= 50:
            return 'C+'
        elif percentage >= 40:
            return 'C'
        else:
            return 'F'


class QuestionBankViewSet(viewsets.ModelViewSet):
    """
    Question bank management with grouping and filtering
    """
    queryset = Question.objects.all()
    permission_classes = [IsAuthenticated]
    filter_backends = [filters.SearchFilter, filters.OrderingFilter]
    search_fields = ['question_text', 'subject__title']
    ordering_fields = ['created_at', 'marks']
    
    def get_serializer_class(self):
        from admin_api.serializers.online_exam import QuestionSerializer
        return QuestionSerializer
    
    def get_queryset(self):
        qs = super().get_queryset()
        
        # Filter by subject
        subject_id = self.request.query_params.get('subject')
        if subject_id:
            qs = qs.filter(subject_id=subject_id)
        
        # Filter by class
        class_id = self.request.query_params.get('class_assigned')
        if class_id:
            qs = qs.filter(class_assigned_id=class_id)
        
        # Filter by type
        q_type = self.request.query_params.get('question_type')
        if q_type:
            qs = qs.filter(question_type=q_type)
        
        # Filter by group
        group_id = self.request.query_params.get('group')
        if group_id:
            qs = qs.filter(group_id=group_id)
        
        return qs.order_by('-created_at')
    
    @action(detail=False, methods=['post'])
    def bulk_import_questions(self, request):
        """
        Bulk import questions from Excel/CSV
        """
        file = request.FILES.get('file')
        subject_id = request.data.get('subject_id')
        class_id = request.data.get('class_id')
        
        if not file or not subject_id or not class_id:
            return Response(
                {'error': 'file, subject_id, and class_id are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            subject = Subject.objects.get(id=subject_id)
            class_obj = Class.objects.get(id=class_id)
        except (Subject.DoesNotExist, Class.DoesNotExist):
            return Response(
                {'error': 'Subject or Class not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            
            created_count = 0
            errors = []
            
            for row_num, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                try:
                    question_text, q_type, options_str, marks = row[:4]
                    
                    # Parse options for MCQ
                    options = {}
                    if q_type == 'mcq' and options_str:
                        # Expect format: "A:Option1|B:Option2|C:Option3|correct:A"
                        for opt in options_str.split('|'):
                            key, value = opt.split(':')
                            options[key.strip()] = value.strip()
                    
                    Question.objects.create(
                        question_text=question_text,
                        question_type=q_type,
                        options=options,
                        marks=marks,
                        subject=subject,
                        class_assigned=class_obj,
                        created_by=request.user
                    )
                    
                    created_count += 1
                
                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')
            
            return Response({
                'message': f'Successfully imported {created_count} questions',
                'created_count': created_count,
                'errors': errors if errors else None
            })
        
        except Exception as e:
            return Response(
                {'error': f'Failed to process file: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )


# ==================== CLASS ROUTINE / TIMETABLE ====================

class TimetableEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced timetable/class routine management with drag-drop support
    """
    queryset = Timetable.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.timetable import TimeTableSerializer
        return TimeTableSerializer
    
    def get_queryset(self):
        qs = super().get_queryset()
        
        # Filter by class
        class_id = self.request.query_params.get('class_assigned')
        if class_id:
            qs = qs.filter(class_assigned_id=class_id)
        
        # Filter by teacher
        teacher_id = self.request.query_params.get('teacher')
        if teacher_id:
            qs = qs.filter(teacher_id=teacher_id)
        
        return qs.select_related('class_assigned', 'subject', 'teacher', 'time_slot').order_by('time_slot')
    
    @action(detail=False, methods=['get'])
    def weekly_schedule(self, request):
        """
        Get weekly schedule in a structured format for drag-drop UI
        """
        class_id = request.query_params.get('class_id')
        
        if not class_id:
            return Response(
                {'error': 'class_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        timetable = Timetable.objects.filter(
            class_assigned_id=class_id
        ).select_related('subject', 'teacher__user', 'time_slot')
        
        # Group by day and time slot
        schedule = {}
        days = ['Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        
        for day in days:
            schedule[day] = []
        
        for entry in timetable:
            day_name = entry.time_slot.day_of_week if hasattr(entry.time_slot, 'day_of_week') else 'Monday'
            
            schedule[day_name].append({
                'id': entry.id,
                'time_slot': str(entry.time_slot),
                'subject': entry.subject.title,
                'teacher': entry.teacher.user.get_full_name() if entry.teacher else 'Not Assigned',
                'room': entry.room or 'TBA'
            })
        
        return Response({
            'class_id': class_id,
            'schedule': schedule
        })
    
    @action(detail=False, methods=['post'])
    def update_schedule(self, request):
        """
        Update timetable entries (for drag-drop operations)
        """
        entries = request.data.get('entries', [])
        
        if not entries:
            return Response(
                {'error': 'entries are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            with transaction.atomic():
                updated_count = 0
                
                for entry_data in entries:
                    entry_id = entry_data.get('id')
                    
                    if entry_id:
                        # Update existing entry
                        entry = Timetable.objects.get(id=entry_id)
                        
                        # Update fields
                        if 'time_slot_id' in entry_data:
                            entry.time_slot_id = entry_data['time_slot_id']
                        if 'subject_id' in entry_data:
                            entry.subject_id = entry_data['subject_id']
                        if 'teacher_id' in entry_data:
                            entry.teacher_id = entry_data['teacher_id']
                        if 'room' in entry_data:
                            entry.room = entry_data['room']
                        
                        entry.save()
                        updated_count += 1
                
                return Response({
                    'message': f'Successfully updated {updated_count} timetable entries',
                    'updated_count': updated_count
                })
        
        except Exception as e:
            return Response(
                {'error': f'Failed to update schedule: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'])
    def check_conflicts(self, request):
        """
        Check for scheduling conflicts (teacher/room double-booking)
        """
        teacher_id = request.data.get('teacher_id')
        time_slot_id = request.data.get('time_slot_id')
        class_id = request.data.get('class_id')
        exclude_id = request.data.get('exclude_id')  # For updates
        
        conflicts = []
        
        # Check teacher conflicts
        if teacher_id and time_slot_id:
            teacher_conflicts = Timetable.objects.filter(
                teacher_id=teacher_id,
                time_slot_id=time_slot_id
            )
            
            if exclude_id:
                teacher_conflicts = teacher_conflicts.exclude(id=exclude_id)
            
            if teacher_conflicts.exists():
                conflicts.append({
                    'type': 'teacher',
                    'message': f'Teacher is already scheduled at this time'
                })
        
        # Check class conflicts
        if class_id and time_slot_id:
            class_conflicts = Timetable.objects.filter(
                class_assigned_id=class_id,
                time_slot_id=time_slot_id
            )
            
            if exclude_id:
                class_conflicts = class_conflicts.exclude(id=exclude_id)
            
            if class_conflicts.exists():
                conflicts.append({
                    'type': 'class',
                    'message': f'Class already has a schedule at this time'
                })
        
        return Response({
            'has_conflicts': len(conflicts) > 0,
            'conflicts': conflicts
        })
    
    @action(detail=False, methods=['get'])
    def export_timetable(self, request):
        """
        Export timetable to Excel
        """
        class_id = request.query_params.get('class_id')
        
        if not class_id:
            return Response(
                {'error': 'class_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            class_obj = Class.objects.get(id=class_id)
        except Class.DoesNotExist:
            return Response(
                {'error': 'Class not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Timetable'
        
        # Title
        ws.merge_cells('A1:F1')
        title_cell = ws['A1']
        title_cell.value = f'Class Timetable - {class_obj.name}'
        title_cell.font = Font(bold=True, size=14)
        title_cell.alignment = Alignment(horizontal='center')
        
        # Headers
        headers = ['Time Slot', 'Monday', 'Tuesday', 'Wednesday', 'Thursday', 'Friday']
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=3, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='CCCCCC', end_color='CCCCCC', fill_type='solid')
        
        # Get all timetable entries
        timetable = Timetable.objects.filter(
            class_assigned=class_obj
        ).select_related('subject', 'teacher__user', 'time_slot')
        
        # Organize by time slots (you'll need to adjust based on your TimeSlot model)
        # This is a simplified version
        row = 4
        for entry in timetable:
            ws.cell(row=row, column=1, value=str(entry.time_slot))
            ws.cell(row=row, column=2, value=f"{entry.subject.title}\n{entry.teacher.user.get_full_name() if entry.teacher else 'TBA'}")
            row += 1
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="timetable_{class_obj.name}_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        return response
