"""
Advanced Academic Module ViewSets
Handles ExamType, ExamMark, ExamResult, GradeScale, Homework, HomeworkSubmission,
LessonPlan, ClassRoutine, StaffAttendance
"""
from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.utils import timezone
from django.db.models import Q, Count, Avg
from datetime import datetime, timedelta
import openpyxl
from io import BytesIO
from django.http import HttpResponse

from admin_api.models import (
    Student, Teacher, Subject, ClassRoom, Exam, ExamResult,
    HomeworkSubmission, LessonPlan, StaffAttendance
)
from admin_api.models_academic import (
    ExamType, ExamMark, GradeScale,
    Homework, ClassRoutine
)
from users.models import User


class ExamTypeViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing exam types (Mid-term, Final, Quiz, etc.)
    """
    queryset = ExamType.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.exam import ExamSerializer
        return ExamSerializer
    
    def get_queryset(self):
        queryset = ExamType.objects.all()
        
        # Filter by active status
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset.order_by('-weightage', 'name')
    
    @action(detail=False, methods=['get'])
    def active_types(self, request):
        """Get all active exam types"""
        from admin_api.serializers.exam import ExamSerializer
        types = ExamType.objects.filter(is_active=True).order_by('name')
        serializer = ExamSerializer(types, many=True)
        return Response(serializer.data)


class ExamMarkViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing exam marks
    """
    queryset = ExamMark.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.exam import ExamResultSerializer
        return ExamResultSerializer
    
    def get_queryset(self):
        queryset = ExamMark.objects.select_related(
            'exam', 'student__user', 'graded_by__user'
        ).all()
        
        # Filter by exam
        exam_id = self.request.query_params.get('exam')
        if exam_id:
            queryset = queryset.filter(exam_id=exam_id)
        
        # Filter by student
        student_id = self.request.query_params.get('student')
        if student_id:
            queryset = queryset.filter(student_id=student_id)
        
        # Filter by absent status
        is_absent = self.request.query_params.get('is_absent')
        if is_absent is not None:
            queryset = queryset.filter(is_absent=is_absent.lower() == 'true')
        
        return queryset.order_by('-exam__exam_date', 'student__roll_no')
    
    def perform_create(self, serializer):
        serializer.save(graded_by=self.request.user.teacher)
    
    def perform_update(self, serializer):
        serializer.save(graded_by=self.request.user.teacher)
    
    @action(detail=False, methods=['post'])
    def bulk_upload(self, request):
        """
        Bulk upload marks from Excel/CSV
        Expects: excel file with columns: roll_no, marks_obtained, is_absent, remarks
        """
        from admin_api.serializers.exam import ExamResultSerializer
        
        exam_id = request.data.get('exam_id')
        file = request.FILES.get('file')
        
        if not exam_id or not file:
            return Response(
                {'error': 'exam_id and file are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            exam = Exam.objects.get(id=exam_id)
        except Exam.DoesNotExist:
            return Response(
                {'error': 'Exam not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        try:
            # Read Excel file
            wb = openpyxl.load_workbook(file)
            sheet = wb.active
            
            created_count = 0
            errors = []
            
            for row_num, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                roll_no, marks, is_absent, remarks = row[:4]
                
                try:
                    student = Student.objects.get(roll_no=roll_no)
                    
                    # Create or update mark
                    mark, created = ExamMark.objects.update_or_create(
                        exam=exam,
                        student=student,
                        defaults={
                            'marks_obtained': marks or 0,
                            'is_absent': bool(is_absent),
                            'remarks': remarks or '',
                            'graded_by': request.user.teacher,
                            'graded_at': timezone.now()
                        }
                    )
                    
                    if created:
                        created_count += 1
                
                except Student.DoesNotExist:
                    errors.append(f'Row {row_num}: Student with roll_no {roll_no} not found')
                except Exception as e:
                    errors.append(f'Row {row_num}: {str(e)}')
            
            return Response({
                'message': f'Successfully uploaded {created_count} marks',
                'created_count': created_count,
                'errors': errors if errors else None
            })
        
        except Exception as e:
            return Response(
                {'error': f'Failed to process file: {str(e)}'},
                status=status.HTTP_400_BAD_REQUEST
            )
    
    @action(detail=False, methods=['get'])
    def export_marks(self, request):
        """
        Export exam marks to Excel
        """
        exam_id = request.query_params.get('exam')
        if not exam_id:
            return Response(
                {'error': 'exam parameter is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            exam = Exam.objects.get(id=exam_id)
        except Exam.DoesNotExist:
            return Response(
                {'error': 'Exam not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Create workbook
        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = 'Exam Marks'
        
        # Headers
        headers = [
            'Roll No', 'Student Name', 'Marks Obtained', 'Total Marks',
            'Percentage', 'Is Absent', 'Grade', 'Remarks', 'Graded By', 'Graded At'
        ]
        sheet.append(headers)
        
        # Fetch marks
        marks = ExamMark.objects.filter(exam=exam).select_related(
            'student__user', 'graded_by__user'
        ).order_by('student__roll_no')
        
        # Add data rows
        for mark in marks:
            percentage = mark.percentage if not mark.is_absent else 0
            sheet.append([
                mark.student.roll_no,
                mark.student.user.get_full_name(),
                float(mark.marks_obtained),
                float(exam.total_marks),
                round(percentage, 2),
                'Yes' if mark.is_absent else 'No',
                mark.get_grade(),
                mark.remarks or '',
                mark.graded_by.user.get_full_name() if mark.graded_by else '',
                mark.graded_at.strftime('%Y-%m-%d %H:%M') if mark.graded_at else ''
            ])
        
        # Save to BytesIO
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        # Return response
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename=exam_marks_{exam.name}_{timezone.now().strftime("%Y%m%d")}.xlsx'
        
        return response


class ExamResultViewSet(viewsets.ModelViewSet):
    """
    ViewSet for consolidated exam results
    """
    queryset = ExamResult.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.exam import ExamResultSerializer
        return ExamResultSerializer
    
    def get_queryset(self):
        queryset = ExamResult.objects.select_related(
            'exam', 'student__user'
        ).all()
        
        # Filter by exam
        exam_id = self.request.query_params.get('exam')
        if exam_id:
            queryset = queryset.filter(exam_id=exam_id)
        
        # Filter by student
        student_id = self.request.query_params.get('student')
        if student_id:
            queryset = queryset.filter(student_id=student_id)
        
        # Filter by passed status
        is_passed = self.request.query_params.get('is_passed')
        if is_passed is not None:
            queryset = queryset.filter(is_passed=is_passed.lower() == 'true')
        
        return queryset.order_by('-exam__exam_date', 'rank')
    
    @action(detail=False, methods=['post'])
    def calculate_results(self, request):
        """
        Calculate and generate results for an exam
        """
        exam_id = request.data.get('exam_id')
        
        if not exam_id:
            return Response(
                {'error': 'exam_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            exam = Exam.objects.get(id=exam_id)
        except Exam.DoesNotExist:
            return Response(
                {'error': 'Exam not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Calculate results from marks
        marks = ExamMark.objects.filter(exam=exam).select_related('student')
        
        results_created = 0
        for mark in marks:
            # Calculate grade based on GradeScale
            percentage = mark.percentage
            grade = self._get_grade(percentage)
            
            # Create or update result
            result, created = ExamResult.objects.update_or_create(
                exam=exam,
                student=mark.student,
                defaults={
                    'total_marks_obtained': mark.marks_obtained,
                    'total_marks': exam.total_marks,
                    'percentage': percentage,
                    'grade': grade,
                    'is_passed': mark.marks_obtained >= exam.passing_marks and not mark.is_absent
                }
            )
            
            if created:
                results_created += 1
        
        # Calculate ranks
        results = ExamResult.objects.filter(exam=exam).order_by('-percentage')
        for rank, result in enumerate(results, start=1):
            result.rank = rank
            result.save(update_fields=['rank'])
        
        return Response({
            'message': f'Results calculated for {results_created} students',
            'total_results': results.count()
        })
    
    def _get_grade(self, percentage):
        """Helper method to get grade from percentage"""
        try:
            grade_scale = GradeScale.objects.filter(
                min_percentage__lte=percentage,
                max_percentage__gte=percentage,
                is_active=True
            ).first()
            return grade_scale.grade if grade_scale else 'N/A'
        except:
            return 'N/A'


class GradeScaleViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing grading scales
    """
    queryset = GradeScale.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.grade import GradeSerializer
        return GradeSerializer
    
    def get_queryset(self):
        queryset = GradeScale.objects.all()
        
        # Filter by active status
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset.order_by('-min_percentage')


class HomeworkViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing homework/assignments
    """
    queryset = Homework.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.assignment import AssignmentSerializer
        return AssignmentSerializer
    
    def get_queryset(self):
        queryset = Homework.objects.select_related(
            'subject', 'class_room', 'assigned_by__user'
        ).all()
        
        # Filter by subject
        subject_id = self.request.query_params.get('subject')
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)
        
        # Filter by class
        class_id = self.request.query_params.get('class_room')
        if class_id:
            queryset = queryset.filter(class_room_id=class_id)
        
        # Filter by status
        homework_status = self.request.query_params.get('status')
        if homework_status:
            queryset = queryset.filter(status=homework_status)
        
        # Filter by assigned teacher
        if self.request.user.role == 'teacher':
            queryset = queryset.filter(assigned_by=self.request.user.teacher)
        
        return queryset.order_by('-assigned_date')
    
    def perform_create(self, serializer):
        serializer.save(assigned_by=self.request.user.teacher)
    
    @action(detail=True, methods=['post'])
    def publish(self, request, pk=None):
        """Publish homework assignment"""
        homework = self.get_object()
        
        if homework.status != 'draft':
            return Response(
                {'error': 'Only draft homework can be published'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        homework.status = 'published'
        homework.save()
        
        # TODO: Send notifications to students
        
        return Response({'message': 'Homework published successfully'})
    
    @action(detail=True, methods=['post'])
    def close(self, request, pk=None):
        """Close homework submissions"""
        homework = self.get_object()
        
        if homework.status == 'closed':
            return Response(
                {'error': 'Homework is already closed'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        homework.status = 'closed'
        homework.save()
        
        return Response({'message': 'Homework closed successfully'})
    
    @action(detail=True, methods=['get'])
    def submissions_summary(self, request, pk=None):
        """Get submission statistics"""
        homework = self.get_object()
        
        total_students = Student.objects.filter(
            class_name=homework.class_room.name,
            is_active=True
        ).count()
        
        submissions = HomeworkSubmission.objects.filter(homework=homework)
        submitted_count = submissions.filter(status__in=['submitted', 'graded', 'late']).count()
        graded_count = submissions.filter(status='graded').count()
        late_count = submissions.filter(status='late').count()
        
        return Response({
            'total_students': total_students,
            'submitted_count': submitted_count,
            'graded_count': graded_count,
            'late_count': late_count,
            'pending_count': total_students - submitted_count,
            'submission_rate': round((submitted_count / total_students * 100), 2) if total_students > 0 else 0
        })


class HomeworkSubmissionViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing homework submissions
    """
    queryset = HomeworkSubmission.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.assignment import AssignmentSerializer
        return AssignmentSerializer
    
    def get_queryset(self):
        queryset = HomeworkSubmission.objects.select_related(
            'homework', 'student__user', 'graded_by__user'
        ).all()
        
        # Filter by homework
        homework_id = self.request.query_params.get('homework')
        if homework_id:
            queryset = queryset.filter(homework_id=homework_id)
        
        # Filter by student
        student_id = self.request.query_params.get('student')
        if student_id:
            queryset = queryset.filter(student_id=student_id)
        elif self.request.user.role == 'student':
            queryset = queryset.filter(student=self.request.user.student)
        
        # Filter by status
        submission_status = self.request.query_params.get('status')
        if submission_status:
            queryset = queryset.filter(status=submission_status)
        
        return queryset.order_by('-submitted_at')
    
    @action(detail=False, methods=['post'])
    def submit(self, request):
        """Submit homework assignment"""
        from admin_api.serializers.assignment import AssignmentSerializer
        
        homework_id = request.data.get('homework_id')
        submission_text = request.data.get('submission_text', '')
        attachment = request.FILES.get('attachment')
        
        if not homework_id:
            return Response(
                {'error': 'homework_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            homework = Homework.objects.get(id=homework_id)
        except Homework.DoesNotExist:
            return Response(
                {'error': 'Homework not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        if homework.status != 'published':
            return Response(
                {'error': 'Homework is not available for submission'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Check if already submitted
        existing_submission = HomeworkSubmission.objects.filter(
            homework=homework,
            student=request.user.student
        ).first()
        
        if existing_submission and existing_submission.status != 'draft':
            return Response(
                {'error': 'You have already submitted this homework'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Determine status (late or submitted)
        submission_status = 'late' if timezone.now().date() > homework.due_date else 'submitted'
        
        # Create or update submission
        submission, created = HomeworkSubmission.objects.update_or_create(
            homework=homework,
            student=request.user.student,
            defaults={
                'submission_text': submission_text,
                'attachment': attachment,
                'submitted_at': timezone.now(),
                'status': submission_status
            }
        )
        
        serializer = AssignmentSerializer(submission)
        return Response(serializer.data, status=status.HTTP_201_CREATED if created else status.HTTP_200_OK)
    
    @action(detail=True, methods=['post'])
    def grade(self, request, pk=None):
        """Grade homework submission"""
        submission = self.get_object()
        
        marks_obtained = request.data.get('marks_obtained')
        feedback = request.data.get('feedback', '')
        
        if marks_obtained is None:
            return Response(
                {'error': 'marks_obtained is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        if float(marks_obtained) > submission.homework.max_marks:
            return Response(
                {'error': 'Marks cannot exceed maximum marks'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        submission.marks_obtained = marks_obtained
        submission.feedback = feedback
        submission.graded_by = request.user.teacher
        submission.graded_at = timezone.now()
        submission.status = 'graded'
        submission.save()
        
        # TODO: Send notification to student
        
        from admin_api.serializers.assignment import AssignmentSerializer
        serializer = AssignmentSerializer(submission)
        return Response(serializer.data)


class LessonPlanViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing lesson plans
    """
    queryset = LessonPlan.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.lesson_plan import LessonPlanSerializer
        return LessonPlanSerializer
    
    def get_queryset(self):
        queryset = LessonPlan.objects.select_related(
            'subject', 'class_room', 'teacher__user', 'academic_year'
        ).all()
        
        # Filter by teacher
        teacher_id = self.request.query_params.get('teacher')
        if teacher_id:
            queryset = queryset.filter(teacher_id=teacher_id)
        elif self.request.user.role == 'teacher':
            queryset = queryset.filter(teacher=self.request.user.teacher)
        
        # Filter by subject
        subject_id = self.request.query_params.get('subject')
        if subject_id:
            queryset = queryset.filter(subject_id=subject_id)
        
        # Filter by class
        class_id = self.request.query_params.get('class_room')
        if class_id:
            queryset = queryset.filter(class_room_id=class_id)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(lesson_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(lesson_date__lte=end_date)
        
        # Filter by completion status
        is_completed = self.request.query_params.get('is_completed')
        if is_completed is not None:
            queryset = queryset.filter(is_completed=is_completed.lower() == 'true')
        
        return queryset.order_by('-lesson_date')
    
    @action(detail=True, methods=['post'])
    def mark_completed(self, request, pk=None):
        """Mark lesson plan as completed"""
        lesson_plan = self.get_object()
        lesson_plan.is_completed = True
        lesson_plan.save()
        
        from admin_api.serializers.lesson_plan import LessonPlanSerializer
        serializer = LessonPlanSerializer(lesson_plan)
        return Response(serializer.data)


class ClassRoutineViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing class timetable/routine
    """
    queryset = ClassRoutine.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.timetable import TimeTableSerializer
        return TimeTableSerializer
    
    def get_queryset(self):
        queryset = ClassRoutine.objects.select_related(
            'class_room', 'subject', 'teacher__user', 'academic_year'
        ).all()
        
        # Filter by class
        class_id = self.request.query_params.get('class_room')
        if class_id:
            queryset = queryset.filter(class_room_id=class_id)
        
        # Filter by teacher
        teacher_id = self.request.query_params.get('teacher')
        if teacher_id:
            queryset = queryset.filter(teacher_id=teacher_id)
        elif self.request.user.role == 'teacher':
            queryset = queryset.filter(teacher=self.request.user.teacher)
        
        # Filter by day
        day = self.request.query_params.get('day_of_week')
        if day:
            queryset = queryset.filter(day_of_week=day)
        
        # Filter by active status
        is_active = self.request.query_params.get('is_active')
        if is_active is not None:
            queryset = queryset.filter(is_active=is_active.lower() == 'true')
        
        return queryset.order_by('day_of_week', 'start_time')
    
    @action(detail=False, methods=['get'])
    def teacher_schedule(self, request):
        """Get teacher's weekly schedule"""
        teacher_id = request.query_params.get('teacher_id')
        
        if not teacher_id:
            if request.user.role == 'teacher':
                teacher_id = request.user.teacher.id
            else:
                return Response(
                    {'error': 'teacher_id is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        routines = ClassRoutine.objects.filter(
            teacher_id=teacher_id,
            is_active=True
        ).select_related('class_room', 'subject').order_by('day_of_week', 'start_time')
        
        # Group by day
        schedule = {}
        for routine in routines:
            day = routine.get_day_of_week_display()
            if day not in schedule:
                schedule[day] = []
            
            from admin_api.serializers.timetable import TimeTableSerializer
            schedule[day].append(TimeTableSerializer(routine).data)
        
        return Response(schedule)
    
    @action(detail=False, methods=['get'])
    def class_schedule(self, request):
        """Get class weekly timetable"""
        class_id = request.query_params.get('class_id')
        
        if not class_id:
            return Response(
                {'error': 'class_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        routines = ClassRoutine.objects.filter(
            class_room_id=class_id,
            is_active=True
        ).select_related('subject', 'teacher__user').order_by('day_of_week', 'start_time')
        
        # Group by day
        schedule = {}
        for routine in routines:
            day = routine.get_day_of_week_display()
            if day not in schedule:
                schedule[day] = []
            
            from admin_api.serializers.timetable import TimeTableSerializer
            schedule[day].append(TimeTableSerializer(routine).data)
        
        return Response(schedule)


class StaffAttendanceViewSet(viewsets.ModelViewSet):
    """
    ViewSet for managing staff/teacher attendance
    """
    queryset = StaffAttendance.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.attendance import AttendanceSerializer
        return AttendanceSerializer
    
    def get_queryset(self):
        queryset = StaffAttendance.objects.select_related(
            'teacher__user', 'marked_by'
        ).all()
        
        # Filter by teacher
        teacher_id = self.request.query_params.get('teacher')
        if teacher_id:
            queryset = queryset.filter(teacher_id=teacher_id)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(date__gte=start_date)
        if end_date:
            queryset = queryset.filter(date__lte=end_date)
        
        # Filter by status
        attendance_status = self.request.query_params.get('status')
        if attendance_status:
            queryset = queryset.filter(status=attendance_status)
        
        return queryset.order_by('-date', 'teacher__user__first_name')
    
    def perform_create(self, serializer):
        serializer.save(marked_by=self.request.user)
    
    @action(detail=False, methods=['post'])
    def mark_attendance(self, request):
        """Bulk mark attendance for multiple teachers"""
        from admin_api.serializers.attendance import AttendanceSerializer
        
        date = request.data.get('date')
        attendance_data = request.data.get('attendance', [])
        
        if not date or not attendance_data:
            return Response(
                {'error': 'date and attendance data are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        created_count = 0
        errors = []
        
        for item in attendance_data:
            teacher_id = item.get('teacher_id')
            status_value = item.get('status', 'present')
            check_in = item.get('check_in_time')
            check_out = item.get('check_out_time')
            remarks = item.get('remarks', '')
            
            try:
                teacher = Teacher.objects.get(id=teacher_id)
                
                # Create or update attendance
                attendance, created = StaffAttendance.objects.update_or_create(
                    teacher=teacher,
                    date=date,
                    defaults={
                        'status': status_value,
                        'check_in_time': check_in,
                        'check_out_time': check_out,
                        'remarks': remarks,
                        'marked_by': request.user
                    }
                )
                
                if created:
                    created_count += 1
            
            except Teacher.DoesNotExist:
                errors.append(f'Teacher with id {teacher_id} not found')
            except Exception as e:
                errors.append(f'Error for teacher {teacher_id}: {str(e)}')
        
        return Response({
            'message': f'Successfully marked attendance for {created_count} teachers',
            'created_count': created_count,
            'errors': errors if errors else None
        })
    
    @action(detail=False, methods=['get'])
    def monthly_report(self, request):
        """Get monthly attendance report"""
        teacher_id = request.query_params.get('teacher_id')
        month = request.query_params.get('month')  # Format: YYYY-MM
        
        if not month:
            return Response(
                {'error': 'month parameter is required (format: YYYY-MM)'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            year, month_num = map(int, month.split('-'))
        except:
            return Response(
                {'error': 'Invalid month format. Use YYYY-MM'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Build query
        queryset = StaffAttendance.objects.filter(
            date__year=year,
            date__month=month_num
        )
        
        if teacher_id:
            queryset = queryset.filter(teacher_id=teacher_id)
        
        # Calculate statistics
        total_days = queryset.count()
        present_days = queryset.filter(status='present').count()
        absent_days = queryset.filter(status='absent').count()
        late_days = queryset.filter(status='late').count()
        half_days = queryset.filter(status='half_day').count()
        on_leave = queryset.filter(status='on_leave').count()
        
        attendance_rate = round((present_days / total_days * 100), 2) if total_days > 0 else 0
        
        return Response({
            'month': month,
            'total_days': total_days,
            'present_days': present_days,
            'absent_days': absent_days,
            'late_days': late_days,
            'half_days': half_days,
            'on_leave': on_leave,
            'attendance_rate': attendance_rate
        })
