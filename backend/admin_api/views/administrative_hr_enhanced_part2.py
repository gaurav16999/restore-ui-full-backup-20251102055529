"""
Administrative & HR Modules - Part 2
=====================================
Fee Management, Wallet System, Inventory, Library, Transport, Dormitory
"""

from rest_framework import viewsets, status
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from django.db.models import Q, Sum, Count, Avg, F
from django.utils import timezone
from django.db import transaction
from django.http import HttpResponse
from datetime import datetime, timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, PatternFill
from io import BytesIO

from admin_api.models import (
    FeeStructure, FeePayment, Student,
    WalletAccount, WalletTransaction, WalletDepositRequest, WalletRefundRequest,
    Supplier, ItemCategory, Item, ItemReceive, ItemIssue,
    BookCategory, LibraryMember, Book, BookIssue,
    TransportRoute, TransportVehicle, VehicleAssignment,
    DormRoomType, DormRoom, DormitoryAssignment,
    User
)


# ==================== FEE MANAGEMENT ====================

class FeeManagementEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced fee management with waivers, refunds, and reporting
    """
    queryset = FeePayment.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.fee import FeePaymentSerializer
        return FeePaymentSerializer
    
    def get_queryset(self):
        queryset = FeePayment.objects.select_related('student__user', 'fee_structure').all()
        
        # Filter by student
        student_id = self.request.query_params.get('student_id')
        if student_id:
            queryset = queryset.filter(student_id=student_id)
        
        # Filter by payment status
        payment_status = self.request.query_params.get('status')
        if payment_status:
            queryset = queryset.filter(status=payment_status)
        
        return queryset.order_by('-payment_date')
    
    @action(detail=False, methods=['post'])
    def apply_waiver(self, request):
        """
        Apply fee waiver to a student
        """
        student_id = request.data.get('student_id')
        fee_structure_id = request.data.get('fee_structure_id')
        waiver_percentage = request.data.get('waiver_percentage', 0)
        waiver_amount = request.data.get('waiver_amount', 0)
        reason = request.data.get('reason', '')
        
        if not student_id or not fee_structure_id:
            return Response(
                {'error': 'student_id and fee_structure_id are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            fee_structure = FeeStructure.objects.get(id=fee_structure_id)
            student = Student.objects.get(id=student_id)
            
            # Calculate waiver
            if waiver_percentage > 0:
                waiver_amount = (fee_structure.amount * Decimal(waiver_percentage)) / Decimal('100')
            
            # Get or create fee payment record
            fee_payment, created = FeePayment.objects.get_or_create(
                student=student,
                fee_structure=fee_structure,
                defaults={
                    'amount_due': fee_structure.amount - Decimal(waiver_amount),
                    'waiver_amount': Decimal(waiver_amount),
                    'waiver_reason': reason,
                    'status': 'partial' if Decimal(waiver_amount) < fee_structure.amount else 'paid'
                }
            )
            
            if not created:
                # Update existing record
                fee_payment.waiver_amount = Decimal(waiver_amount)
                fee_payment.waiver_reason = reason
                fee_payment.amount_due = fee_structure.amount - Decimal(waiver_amount) - fee_payment.amount_paid
                fee_payment.save()
            
            return Response({
                'message': 'Fee waiver applied successfully',
                'student': student.user.get_full_name(),
                'original_amount': float(fee_structure.amount),
                'waiver_amount': float(waiver_amount),
                'amount_due': float(fee_payment.amount_due)
            })
        
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['post'])
    def process_refund(self, request, pk=None):
        """
        Process fee refund
        """
        fee_payment = self.get_object()
        refund_amount = request.data.get('refund_amount')
        reason = request.data.get('reason', '')
        
        if not refund_amount:
            return Response(
                {'error': 'refund_amount is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        refund_amount = Decimal(refund_amount)
        
        if refund_amount > fee_payment.amount_paid:
            return Response(
                {'error': 'Refund amount cannot exceed paid amount'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            with transaction.atomic():
                # Update fee payment
                fee_payment.amount_paid -= refund_amount
                fee_payment.amount_due += refund_amount
                fee_payment.refund_amount = refund_amount
                fee_payment.refund_reason = reason
                fee_payment.refund_date = timezone.now()
                
                # Update status
                if fee_payment.amount_paid == 0:
                    fee_payment.status = 'pending'
                elif fee_payment.amount_paid < fee_payment.amount_due:
                    fee_payment.status = 'partial'
                
                fee_payment.save()
                
                return Response({
                    'message': 'Refund processed successfully',
                    'refund_amount': float(refund_amount),
                    'remaining_balance': float(fee_payment.amount_paid),
                    'amount_due': float(fee_payment.amount_due)
                })
        
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def collection_report(self, request):
        """
        Generate fee collection report
        """
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date', timezone.now().date())
        
        queryset = FeePayment.objects.all()
        
        if start_date:
            queryset = queryset.filter(payment_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(payment_date__lte=end_date)
        
        summary = queryset.aggregate(
            total_collected=Sum('amount_paid'),
            total_due=Sum('amount_due'),
            total_waiver=Sum('waiver_amount'),
            total_refund=Sum('refund_amount'),
            total_students=Count('student', distinct=True)
        )
        
        # Get status-wise breakdown
        status_breakdown = queryset.values('status').annotate(
            count=Count('id'),
            total_amount=Sum('amount_paid')
        )
        
        return Response({
            'period': f'{start_date} to {end_date}',
            'summary': {
                'total_collected': float(summary['total_collected'] or 0),
                'total_due': float(summary['total_due'] or 0),
                'total_waiver': float(summary['total_waiver'] or 0),
                'total_refund': float(summary['total_refund'] or 0),
                'total_students': summary['total_students'] or 0
            },
            'status_breakdown': list(status_breakdown)
        })
    
    @action(detail=False, methods=['get'])
    def outstanding_fees(self, request):
        """
        Get list of students with outstanding fees
        """
        outstanding = FeePayment.objects.filter(
            status__in=['pending', 'partial']
        ).select_related('student__user', 'fee_structure').order_by('-amount_due')
        
        data = []
        for payment in outstanding:
            data.append({
                'student_id': payment.student.id,
                'student_name': payment.student.user.get_full_name(),
                'class': payment.student.current_class.name if payment.student.current_class else '',
                'fee_type': payment.fee_structure.name,
                'total_amount': float(payment.fee_structure.amount),
                'amount_paid': float(payment.amount_paid),
                'amount_due': float(payment.amount_due),
                'due_date': payment.due_date,
                'status': payment.status
            })
        
        total_outstanding = sum(item['amount_due'] for item in data)
        
        return Response({
            'total_outstanding': total_outstanding,
            'student_count': len(data),
            'outstanding_fees': data
        })
    
    @action(detail=False, methods=['get'])
    def export_fee_report(self, request):
        """
        Export fee collection report to Excel
        """
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date', timezone.now().date())
        
        queryset = FeePayment.objects.all()
        
        if start_date:
            queryset = queryset.filter(payment_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(payment_date__lte=end_date)
        
        queryset = queryset.select_related('student__user', 'fee_structure')
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Fee Collection Report'
        
        # Header
        headers = [
            'Student ID', 'Student Name', 'Class', 'Fee Type',
            'Total Amount', 'Amount Paid', 'Amount Due', 'Waiver',
            'Payment Date', 'Status'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data rows
        for row_num, payment in enumerate(queryset, start=2):
            ws.cell(row=row_num, column=1, value=payment.student.student_id if hasattr(payment.student, 'student_id') else '')
            ws.cell(row=row_num, column=2, value=payment.student.user.get_full_name())
            ws.cell(row=row_num, column=3, value=payment.student.current_class.name if payment.student.current_class else '')
            ws.cell(row=row_num, column=4, value=payment.fee_structure.name)
            ws.cell(row=row_num, column=5, value=float(payment.fee_structure.amount))
            ws.cell(row=row_num, column=6, value=float(payment.amount_paid))
            ws.cell(row=row_num, column=7, value=float(payment.amount_due))
            ws.cell(row=row_num, column=8, value=float(payment.waiver_amount or 0))
            ws.cell(row=row_num, column=9, value=payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else '')
            ws.cell(row=row_num, column=10, value=payment.status.title())
        
        # Save
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="fee_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        return response


# ==================== WALLET SYSTEM ====================

class WalletSystemEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced wallet system with transfers and transaction history
    """
    queryset = WalletAccount.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.wallet import WalletAccountSerializer
        return WalletAccountSerializer
    
    @action(detail=False, methods=['post'])
    def transfer_funds(self, request):
        """
        Transfer funds between wallets
        """
        from_wallet_id = request.data.get('from_wallet_id')
        to_wallet_id = request.data.get('to_wallet_id')
        amount = request.data.get('amount')
        description = request.data.get('description', 'Wallet Transfer')
        
        if not from_wallet_id or not to_wallet_id or not amount:
            return Response(
                {'error': 'from_wallet_id, to_wallet_id, and amount are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        amount = Decimal(amount)
        
        try:
            with transaction.atomic():
                from_wallet = WalletAccount.objects.select_for_update().get(id=from_wallet_id)
                to_wallet = WalletAccount.objects.select_for_update().get(id=to_wallet_id)
                
                # Check balance
                if from_wallet.balance < amount:
                    return Response(
                        {'error': 'Insufficient balance'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Deduct from source
                from_wallet.balance -= amount
                from_wallet.save()
                
                # Add to destination
                to_wallet.balance += amount
                to_wallet.save()
                
                # Create transactions
                WalletTransaction.objects.create(
                    wallet=from_wallet,
                    transaction_type='debit',
                    amount=amount,
                    description=f'{description} - To {to_wallet.student.user.get_full_name() if hasattr(to_wallet, "student") else "Unknown"}',
                    balance_after=from_wallet.balance
                )
                
                WalletTransaction.objects.create(
                    wallet=to_wallet,
                    transaction_type='credit',
                    amount=amount,
                    description=f'{description} - From {from_wallet.student.user.get_full_name() if hasattr(from_wallet, "student") else "Unknown"}',
                    balance_after=to_wallet.balance
                )
                
                return Response({
                    'message': 'Transfer completed successfully',
                    'from_wallet_balance': float(from_wallet.balance),
                    'to_wallet_balance': float(to_wallet.balance),
                    'amount_transferred': float(amount)
                })
        
        except WalletAccount.DoesNotExist:
            return Response(
                {'error': 'Wallet not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        except Exception as e:
            return Response(
                {'error': str(e)},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=True, methods=['get'])
    def transaction_history(self, request, pk=None):
        """
        Get transaction history for a wallet
        """
        wallet = self.get_object()
        
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        transactions = WalletTransaction.objects.filter(wallet=wallet)
        
        if start_date:
            transactions = transactions.filter(created_at__gte=start_date)
        if end_date:
            transactions = transactions.filter(created_at__lte=end_date)
        
        transactions = transactions.order_by('-created_at')
        
        data = []
        for txn in transactions:
            data.append({
                'id': txn.id,
                'date': txn.created_at,
                'type': txn.transaction_type,
                'amount': float(txn.amount),
                'description': txn.description,
                'balance_after': float(txn.balance_after)
            })
        
        return Response({
            'wallet_id': wallet.id,
            'current_balance': float(wallet.balance),
            'transaction_count': len(data),
            'transactions': data
        })
    
    @action(detail=False, methods=['post'])
    def deposit(self, request):
        """
        Create deposit request
        """
        wallet_id = request.data.get('wallet_id')
        amount = request.data.get('amount')
        payment_method = request.data.get('payment_method', 'cash')
        
        if not wallet_id or not amount:
            return Response(
                {'error': 'wallet_id and amount are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            wallet = WalletAccount.objects.get(id=wallet_id)
            
            deposit_request = WalletDepositRequest.objects.create(
                wallet=wallet,
                amount=Decimal(amount),
                payment_method=payment_method,
                requested_by=request.user
            )
            
            return Response({
                'message': 'Deposit request created successfully',
                'request_id': deposit_request.id,
                'amount': float(amount),
                'status': 'pending'
            })
        
        except WalletAccount.DoesNotExist:
            return Response(
                {'error': 'Wallet not found'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=False, methods=['post'])
    def approve_deposit(self, request):
        """
        Approve deposit request and credit wallet
        """
        request_id = request.data.get('request_id')
        
        if not request_id:
            return Response(
                {'error': 'request_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            with transaction.atomic():
                deposit_req = WalletDepositRequest.objects.select_for_update().get(id=request_id)
                
                if deposit_req.status != 'pending':
                    return Response(
                        {'error': 'Deposit request already processed'},
                        status=status.HTTP_400_BAD_REQUEST
                    )
                
                # Update wallet balance
                wallet = deposit_req.wallet
                wallet.balance += deposit_req.amount
                wallet.save()
                
                # Update deposit request
                deposit_req.status = 'approved'
                deposit_req.approved_by = request.user
                deposit_req.approved_at = timezone.now()
                deposit_req.save()
                
                # Create transaction
                WalletTransaction.objects.create(
                    wallet=wallet,
                    transaction_type='credit',
                    amount=deposit_req.amount,
                    description='Deposit approved',
                    balance_after=wallet.balance
                )
                
                return Response({
                    'message': 'Deposit approved successfully',
                    'amount': float(deposit_req.amount),
                    'new_balance': float(wallet.balance)
                })
        
        except WalletDepositRequest.DoesNotExist:
            return Response(
                {'error': 'Deposit request not found'},
                status=status.HTTP_404_NOT_FOUND
            )
    
    @action(detail=False, methods=['get'])
    def balance_report(self, request):
        """
        Get wallet balance report for all accounts
        """
        wallets = WalletAccount.objects.all()
        
        data = []
        total_balance = Decimal('0.00')
        
        for wallet in wallets:
            balance = wallet.balance
            total_balance += balance
            
            data.append({
                'wallet_id': wallet.id,
                'owner': wallet.student.user.get_full_name() if hasattr(wallet, 'student') else 'Unknown',
                'balance': float(balance),
                'last_transaction': WalletTransaction.objects.filter(wallet=wallet).order_by('-created_at').first().created_at if WalletTransaction.objects.filter(wallet=wallet).exists() else None
            })
        
        return Response({
            'total_wallets': wallets.count(),
            'total_balance': float(total_balance),
            'wallets': data
        })


# ==================== INVENTORY MANAGEMENT ====================

class InventoryEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced inventory with stock alerts and reports
    """
    queryset = Item.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.inventory import ItemSerializer
        return ItemSerializer
    
    @action(detail=False, methods=['get'])
    def low_stock_alert(self, request):
        """
        Get items with low stock
        """
        threshold = int(request.query_params.get('threshold', 10))
        
        low_stock_items = Item.objects.filter(
            quantity__lte=threshold,
            is_active=True
        ).select_related('category', 'supplier')
        
        data = []
        for item in low_stock_items:
            data.append({
                'id': item.id,
                'name': item.name,
                'category': item.category.name if item.category else '',
                'current_quantity': item.quantity,
                'unit': item.unit,
                'supplier': item.supplier.name if item.supplier else '',
                'reorder_level': threshold
            })
        
        return Response({
            'threshold': threshold,
            'low_stock_count': len(data),
            'items': data
        })
    
    @action(detail=False, methods=['get'])
    def stock_movement_report(self, request):
        """
        Generate stock movement report
        """
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date', timezone.now().date())
        
        # Get receives
        receives = ItemReceive.objects.all()
        if start_date:
            receives = receives.filter(receive_date__gte=start_date)
        if end_date:
            receives = receives.filter(receive_date__lte=end_date)
        
        total_received = receives.aggregate(
            total_quantity=Sum('quantity'),
            total_value=Sum(F('quantity') * F('unit_price'))
        )
        
        # Get issues
        issues = ItemIssue.objects.all()
        if start_date:
            issues = issues.filter(issue_date__gte=start_date)
        if end_date:
            issues = issues.filter(issue_date__lte=end_date)
        
        total_issued = issues.aggregate(
            total_quantity=Sum('quantity')
        )
        
        return Response({
            'period': f'{start_date} to {end_date}',
            'received': {
                'total_quantity': total_received['total_quantity'] or 0,
                'total_value': float(total_received['total_value'] or 0),
                'transaction_count': receives.count()
            },
            'issued': {
                'total_quantity': total_issued['total_quantity'] or 0,
                'transaction_count': issues.count()
            }
        })
    
    @action(detail=False, methods=['get'])
    def inventory_valuation(self, request):
        """
        Calculate total inventory value
        """
        items = Item.objects.filter(is_active=True)
        
        total_value = Decimal('0.00')
        item_data = []
        
        for item in items:
            # Get latest unit price from ItemReceive
            latest_receive = ItemReceive.objects.filter(item=item).order_by('-receive_date').first()
            unit_price = latest_receive.unit_price if latest_receive else Decimal('0.00')
            
            item_value = item.quantity * unit_price
            total_value += item_value
            
            item_data.append({
                'id': item.id,
                'name': item.name,
                'quantity': item.quantity,
                'unit_price': float(unit_price),
                'total_value': float(item_value)
            })
        
        return Response({
            'total_inventory_value': float(total_value),
            'total_items': items.count(),
            'items': item_data
        })


# ==================== LIBRARY MANAGEMENT ====================

class LibraryEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced library with overdue tracking and statistics
    """
    queryset = Book.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.library import BookSerializer
        return BookSerializer
    
    @action(detail=False, methods=['get'])
    def overdue_books(self, request):
        """
        Get list of overdue books with fine calculation
        """
        today = timezone.now().date()
        
        overdue_issues = BookIssue.objects.filter(
            return_date__isnull=True,
            due_date__lt=today
        ).select_related('book', 'member')
        
        fine_per_day = Decimal('5.00')  # Configure as needed
        
        data = []
        total_fine = Decimal('0.00')
        
        for issue in overdue_issues:
            days_overdue = (today - issue.due_date).days
            fine = days_overdue * fine_per_day
            total_fine += fine
            
            data.append({
                'issue_id': issue.id,
                'book_title': issue.book.title,
                'member_name': issue.member.name if hasattr(issue.member, 'name') else '',
                'issue_date': issue.issue_date,
                'due_date': issue.due_date,
                'days_overdue': days_overdue,
                'fine_amount': float(fine)
            })
        
        return Response({
            'overdue_count': len(data),
            'total_fine': float(total_fine),
            'fine_per_day': float(fine_per_day),
            'overdue_books': data
        })
    
    @action(detail=False, methods=['get'])
    def issue_statistics(self, request):
        """
        Get book issue statistics
        """
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date', timezone.now().date())
        
        issues = BookIssue.objects.all()
        
        if start_date:
            issues = issues.filter(issue_date__gte=start_date)
        if end_date:
            issues = issues.filter(issue_date__lte=end_date)
        
        total_issued = issues.count()
        total_returned = issues.filter(return_date__isnull=False).count()
        currently_issued = issues.filter(return_date__isnull=True).count()
        
        return Response({
            'period': f'{start_date} to {end_date}',
            'total_issued': total_issued,
            'total_returned': total_returned,
            'currently_issued': currently_issued,
            'return_rate': round((total_returned / total_issued * 100), 2) if total_issued > 0 else 0
        })
    
    @action(detail=False, methods=['get'])
    def popular_books(self, request):
        """
        Get most issued books
        """
        limit = int(request.query_params.get('limit', 10))
        
        popular = BookIssue.objects.values('book__title', 'book__author').annotate(
            issue_count=Count('id')
        ).order_by('-issue_count')[:limit]
        
        return Response({
            'popular_books': list(popular)
        })


# ==================== TRANSPORT MANAGEMENT ====================

class TransportEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced transport with route-wise student tracking
    """
    queryset = TransportRoute.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.transport import TransportRouteSerializer
        return TransportRouteSerializer
    
    @action(detail=True, methods=['get'])
    def route_students(self, request, pk=None):
        """
        Get all students assigned to a route
        """
        route = self.get_object()
        
        assignments = VehicleAssignment.objects.filter(
            route=route
        ).select_related('student__user', 'vehicle')
        
        data = []
        for assignment in assignments:
            data.append({
                'student_id': assignment.student.id,
                'student_name': assignment.student.user.get_full_name(),
                'class': assignment.student.current_class.name if assignment.student.current_class else '',
                'vehicle': assignment.vehicle.vehicle_number if assignment.vehicle else '',
                'pickup_point': assignment.pickup_point if hasattr(assignment, 'pickup_point') else ''
            })
        
        return Response({
            'route_name': route.route_name,
            'student_count': len(data),
            'students': data
        })
    
    @action(detail=False, methods=['get'])
    def vehicle_summary(self, request):
        """
        Get summary of all vehicles and capacity
        """
        vehicles = TransportVehicle.objects.all()
        
        data = []
        total_capacity = 0
        total_assigned = 0
        
        for vehicle in vehicles:
            assigned_count = VehicleAssignment.objects.filter(vehicle=vehicle).count()
            
            data.append({
                'vehicle_number': vehicle.vehicle_number,
                'capacity': vehicle.capacity if hasattr(vehicle, 'capacity') else 0,
                'assigned': assigned_count,
                'available': (vehicle.capacity if hasattr(vehicle, 'capacity') else 0) - assigned_count
            })
            
            if hasattr(vehicle, 'capacity'):
                total_capacity += vehicle.capacity
            total_assigned += assigned_count
        
        return Response({
            'total_vehicles': vehicles.count(),
            'total_capacity': total_capacity,
            'total_assigned': total_assigned,
            'available_seats': total_capacity - total_assigned,
            'vehicles': data
        })


# ==================== DORMITORY MANAGEMENT ====================

class DormitoryEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced dormitory with occupancy tracking
    """
    queryset = DormRoom.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.dormitory import DormRoomSerializer
        return DormRoomSerializer
    
    @action(detail=False, methods=['get'])
    def occupancy_report(self, request):
        """
        Get dormitory occupancy report
        """
        rooms = DormRoom.objects.all()
        
        data = []
        total_capacity = 0
        total_occupied = 0
        
        for room in rooms:
            occupied = DormitoryAssignment.objects.filter(room=room, is_active=True).count()
            capacity = room.capacity if hasattr(room, 'capacity') else 0
            
            data.append({
                'room_number': room.room_number,
                'room_type': room.room_type.name if hasattr(room, 'room_type') else '',
                'capacity': capacity,
                'occupied': occupied,
                'available': capacity - occupied,
                'occupancy_rate': round((occupied / capacity * 100), 2) if capacity > 0 else 0
            })
            
            total_capacity += capacity
            total_occupied += occupied
        
        return Response({
            'total_rooms': rooms.count(),
            'total_capacity': total_capacity,
            'total_occupied': total_occupied,
            'total_available': total_capacity - total_occupied,
            'overall_occupancy_rate': round((total_occupied / total_capacity * 100), 2) if total_capacity > 0 else 0,
            'rooms': data
        })
    
    @action(detail=True, methods=['get'])
    def room_assignments(self, request, pk=None):
        """
        Get all students assigned to a room
        """
        room = self.get_object()
        
        assignments = DormitoryAssignment.objects.filter(
            room=room,
            is_active=True
        ).select_related('student__user')
        
        data = []
        for assignment in assignments:
            data.append({
                'student_id': assignment.student.id,
                'student_name': assignment.student.user.get_full_name(),
                'class': assignment.student.current_class.name if assignment.student.current_class else '',
                'assigned_date': assignment.assigned_date if hasattr(assignment, 'assigned_date') else None
            })
        
        return Response({
            'room_number': room.room_number,
            'capacity': room.capacity if hasattr(room, 'capacity') else 0,
            'occupied': len(data),
            'students': data
        })
