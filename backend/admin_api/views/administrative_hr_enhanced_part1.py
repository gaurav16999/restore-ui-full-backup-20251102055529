"""
Administrative & HR Modules - Complete Enhancement
====================================================
Comprehensive enhancements for:
1. HR & Payroll (Leave Types, Applications, Approvals, Payroll Runs)
2. Staff Attendance (Daily + Monthly Summary with Reports)
3. Double-Entry Accounting (Chart of Accounts → Ledgers → Vouchers → Trial Balance → Reports)
4. Fee Management (Structures, Waivers, Online Payments, Refunds)
5. Wallet System (Deposit, Refund, Transfer, Transaction History)
6. Inventory, Library, Transport, Dormitory (End-to-end CRUD + Reporting)
"""

from rest_framework import viewsets, status, filters
from rest_framework.decorators import action
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated
from rest_framework.parsers import MultiPartParser, FormParser
from django.db.models import Q, Avg, Count, Sum, Max, Min, F
from django.utils import timezone
from django.db import transaction
from django.http import HttpResponse
from datetime import datetime, timedelta
from decimal import Decimal
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from io import BytesIO
import csv

from admin_api.models import (
    Teacher, Employee, Designation, Department,
    StaffAttendance, PayrollRecord,
    LeaveType, LeaveDefine, LeaveApplication,
    ChartOfAccount, AccountTransaction, AccountGroup, JournalEntry, JournalEntryLine, BudgetAllocation,
    FeeStructure, FeePayment, Student,
    WalletAccount, WalletTransaction, WalletDepositRequest, WalletRefundRequest,
    Supplier, ItemCategory, Item, ItemReceive, ItemIssue,
    BookCategory, LibraryMember, Book, BookIssue,
    TransportRoute, TransportVehicle, VehicleAssignment,
    DormRoomType, DormRoom, DormitoryAssignment,
    User
)


# ==================== HR & PAYROLL ====================

class LeaveManagementEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced leave management with approvals and balance tracking
    """
    queryset = LeaveApplication.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.leave import LeaveApplicationSerializer
        return LeaveApplicationSerializer
    
    def get_queryset(self):
        queryset = LeaveApplication.objects.select_related('teacher__user', 'leave_type').all()
        
        # Filter by status
        leave_status = self.request.query_params.get('status')
        if leave_status:
            queryset = queryset.filter(status=leave_status)
        
        # Filter by teacher (for teachers viewing their own)
        if self.request.user.role == 'teacher':
            queryset = queryset.filter(teacher=self.request.user.teacher)
        
        # Filter by leave type
        leave_type_id = self.request.query_params.get('leave_type')
        if leave_type_id:
            queryset = queryset.filter(leave_type_id=leave_type_id)
        
        # Filter by date range
        start_date = self.request.query_params.get('start_date')
        end_date = self.request.query_params.get('end_date')
        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(end_date__lte=end_date)
        
        return queryset.order_by('-applied_date')
    
    @action(detail=True, methods=['post'])
    def approve(self, request, pk=None):
        """
        Approve leave application
        """
        leave_app = self.get_object()
        
        if leave_app.status != 'pending':
            return Response(
                {'error': 'Only pending applications can be approved'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        leave_app.status = 'approved'
        leave_app.approved_by = request.user
        leave_app.approved_date = timezone.now().date()
        leave_app.save()
        
        return Response({
            'message': 'Leave application approved successfully',
            'data': self.get_serializer(leave_app).data
        })
    
    @action(detail=True, methods=['post'])
    def reject(self, request, pk=None):
        """
        Reject leave application
        """
        leave_app = self.get_object()
        rejection_reason = request.data.get('reason', '')
        
        if leave_app.status != 'pending':
            return Response(
                {'error': 'Only pending applications can be rejected'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        leave_app.status = 'rejected'
        leave_app.approved_by = request.user
        leave_app.approved_date = timezone.now().date()
        leave_app.remarks = rejection_reason
        leave_app.save()
        
        return Response({
            'message': 'Leave application rejected',
            'data': self.get_serializer(leave_app).data
        })
    
    @action(detail=False, methods=['get'])
    def leave_balance(self, request):
        """
        Get leave balance for a teacher
        """
        teacher_id = request.query_params.get('teacher_id')
        
        if not teacher_id:
            if request.user.role == 'teacher':
                teacher_id = request.user.teacher.id
            else:
                return Response(
                    {'error': 'teacher_id is required'},
                    status=status.HTTP_400_BAD_REQUEST
                )
        
        # Get all leave types
        leave_types = LeaveType.objects.all()
        
        balance_data = []
        for leave_type in leave_types:
            # Get annual allocation
            leave_define = LeaveDefine.objects.filter(
                leave_type=leave_type
            ).first()
            
            annual_allocation = leave_define.days if leave_define else 0
            
            # Calculate used leaves (approved only)
            used_leaves = LeaveApplication.objects.filter(
                teacher_id=teacher_id,
                leave_type=leave_type,
                status='approved',
                start_date__year=timezone.now().year
            ).aggregate(
                total_days=Sum(F('end_date') - F('start_date'))
            )['total_days'] or 0
            
            # Convert timedelta to days
            if hasattr(used_leaves, 'days'):
                used_leaves = used_leaves.days + 1  # Include end date
            
            balance_data.append({
                'leave_type': leave_type.name,
                'annual_allocation': annual_allocation,
                'used': used_leaves,
                'balance': annual_allocation - used_leaves
            })
        
        return Response({
            'teacher_id': teacher_id,
            'year': timezone.now().year,
            'leave_balance': balance_data
        })
    
    @action(detail=False, methods=['get'])
    def pending_approvals(self, request):
        """
        Get all pending leave applications for approval
        """
        pending = LeaveApplication.objects.filter(
            status='pending'
        ).select_related('teacher__user', 'leave_type').order_by('applied_date')
        
        serializer = self.get_serializer(pending, many=True)
        
        return Response({
            'total_pending': pending.count(),
            'applications': serializer.data
        })
    
    @action(detail=False, methods=['get'])
    def export_leave_report(self, request):
        """
        Export leave report to Excel
        """
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        queryset = self.get_queryset()
        
        if start_date:
            queryset = queryset.filter(start_date__gte=start_date)
        if end_date:
            queryset = queryset.filter(end_date__lte=end_date)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Leave Report'
        
        # Header
        headers = [
            'Employee ID', 'Employee Name', 'Leave Type', 'Start Date', 'End Date',
            'Days', 'Status', 'Applied Date', 'Approved By', 'Reason'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data rows
        for row_num, leave_app in enumerate(queryset, start=2):
            days = (leave_app.end_date - leave_app.start_date).days + 1
            
            ws.cell(row=row_num, column=1, value=leave_app.teacher.employee_id if hasattr(leave_app.teacher, 'employee_id') else '')
            ws.cell(row=row_num, column=2, value=leave_app.teacher.user.get_full_name())
            ws.cell(row=row_num, column=3, value=leave_app.leave_type.name)
            ws.cell(row=row_num, column=4, value=leave_app.start_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=5, value=leave_app.end_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=6, value=days)
            ws.cell(row=row_num, column=7, value=leave_app.status.title())
            ws.cell(row=row_num, column=8, value=leave_app.applied_date.strftime('%Y-%m-%d'))
            ws.cell(row=row_num, column=9, value=leave_app.approved_by.get_full_name() if leave_app.approved_by else '')
            ws.cell(row=row_num, column=10, value=leave_app.reason or '')
        
        # Save
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="leave_report_{timezone.now().strftime("%Y%m%d")}.xlsx"'
        
        return response


class PayrollRunEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced payroll processing with detailed calculations
    """
    queryset = PayrollRecord.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.hr import PayrollRecordSerializer
        return PayrollRecordSerializer
    
    def get_queryset(self):
        queryset = PayrollRecord.objects.select_related('teacher__user').all()
        
        # Filter by month/year
        month = self.request.query_params.get('month')
        year = self.request.query_params.get('year')
        if month:
            queryset = queryset.filter(month=int(month))
        if year:
            queryset = queryset.filter(year=int(year))
        
        return queryset.order_by('-year', '-month', 'teacher__user__first_name')
    
    @action(detail=False, methods=['post'])
    def generate_payroll(self, request):
        """
        Generate payroll for a specific month
        """
        month = request.data.get('month')
        year = request.data.get('year')
        
        if not month or not year:
            return Response(
                {'error': 'month and year are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        month = int(month)
        year = int(year)
        
        # Check if payroll already exists
        existing = PayrollRecord.objects.filter(month=month, year=year).count()
        if existing > 0:
            return Response(
                {'error': f'Payroll for {month}/{year} already exists. Please delete or modify existing records.'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Get all active teachers
        teachers = Teacher.objects.filter(is_active=True)
        
        created_count = 0
        
        with transaction.atomic():
            for teacher in teachers:
                # Calculate salary components
                basic_salary = teacher.salary or Decimal('0.00')
                allowances = Decimal('0.00')  # Add logic for allowances
                deductions = Decimal('0.00')  # Add logic for deductions
                
                # Calculate attendance-based deductions
                working_days = 30  # Adjust based on month
                attendance_count = StaffAttendance.objects.filter(
                    teacher=teacher,
                    date__month=month,
                    date__year=year,
                    status='present'
                ).count()
                
                # Absence deduction
                absent_days = working_days - attendance_count
                per_day_salary = basic_salary / Decimal(working_days)
                absence_deduction = per_day_salary * Decimal(absent_days)
                deductions += absence_deduction
                
                # Calculate net salary
                gross_salary = basic_salary + allowances
                net_salary = gross_salary - deductions
                
                # Create payroll record
                PayrollRecord.objects.create(
                    teacher=teacher,
                    month=month,
                    year=year,
                    basic_salary=basic_salary,
                    allowances=allowances,
                    deductions=deductions,
                    gross_salary=gross_salary,
                    net_salary=net_salary
                )
                
                created_count += 1
        
        return Response({
            'message': f'Payroll generated successfully for {month}/{year}',
            'records_created': created_count,
            'month': month,
            'year': year
        })
    
    @action(detail=False, methods=['get'])
    def payroll_summary(self, request):
        """
        Get payroll summary for a month
        """
        month = request.query_params.get('month')
        year = request.query_params.get('year')
        
        if not month or not year:
            return Response(
                {'error': 'month and year are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        records = PayrollRecord.objects.filter(month=int(month), year=int(year))
        
        summary = records.aggregate(
            total_employees=Count('id'),
            total_gross=Sum('gross_salary'),
            total_deductions=Sum('deductions'),
            total_net=Sum('net_salary'),
            avg_salary=Avg('net_salary')
        )
        
        return Response({
            'month': month,
            'year': year,
            'total_employees': summary['total_employees'] or 0,
            'total_gross_salary': float(summary['total_gross'] or 0),
            'total_deductions': float(summary['total_deductions'] or 0),
            'total_net_salary': float(summary['total_net'] or 0),
            'average_salary': float(summary['avg_salary'] or 0)
        })
    
    @action(detail=False, methods=['get'])
    def export_payslips(self, request):
        """
        Export payslips for all employees
        """
        month = request.query_params.get('month')
        year = request.query_params.get('year')
        
        if not month or not year:
            return Response(
                {'error': 'month and year are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        records = PayrollRecord.objects.filter(
            month=int(month),
            year=int(year)
        ).select_related('teacher__user')
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Payslips'
        
        # Header
        headers = [
            'Employee ID', 'Employee Name', 'Basic Salary', 'Allowances',
            'Gross Salary', 'Deductions', 'Net Salary', 'Payment Status'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data rows
        for row_num, record in enumerate(records, start=2):
            ws.cell(row=row_num, column=1, value=record.teacher.employee_id if hasattr(record.teacher, 'employee_id') else '')
            ws.cell(row=row_num, column=2, value=record.teacher.user.get_full_name())
            ws.cell(row=row_num, column=3, value=float(record.basic_salary))
            ws.cell(row=row_num, column=4, value=float(record.allowances))
            ws.cell(row=row_num, column=5, value=float(record.gross_salary))
            ws.cell(row=row_num, column=6, value=float(record.deductions))
            ws.cell(row=row_num, column=7, value=float(record.net_salary))
            ws.cell(row=row_num, column=8, value='Paid' if record.is_paid else 'Pending')
        
        # Save
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="payslips_{month}_{year}.xlsx"'
        
        return response


# ==================== STAFF ATTENDANCE ====================

class StaffAttendanceEnhancedViewSet(viewsets.ModelViewSet):
    """
    Enhanced staff attendance with daily and monthly reporting
    """
    queryset = StaffAttendance.objects.all()
    permission_classes = [IsAuthenticated]
    
    def get_serializer_class(self):
        from admin_api.serializers.hr import StaffAttendanceSerializer
        return StaffAttendanceSerializer
    
    def get_queryset(self):
        queryset = StaffAttendance.objects.select_related('teacher__user').all()
        
        # Filter by date
        date = self.request.query_params.get('date')
        if date:
            queryset = queryset.filter(date=date)
        
        # Filter by teacher
        teacher_id = self.request.query_params.get('teacher_id')
        if teacher_id:
            queryset = queryset.filter(teacher_id=teacher_id)
        
        # Filter by status
        attendance_status = self.request.query_params.get('status')
        if attendance_status:
            queryset = queryset.filter(status=attendance_status)
        
        return queryset.order_by('-date', 'teacher__user__first_name')
    
    @action(detail=False, methods=['post'])
    def mark_daily_attendance(self, request):
        """
        Mark attendance for multiple staff members
        """
        date = request.data.get('date')
        attendance_records = request.data.get('attendance', [])
        
        if not date or not attendance_records:
            return Response(
                {'error': 'date and attendance records are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        created_count = 0
        updated_count = 0
        
        with transaction.atomic():
            for record in attendance_records:
                teacher_id = record.get('teacher_id')
                attendance_status = record.get('status', 'present')
                remarks = record.get('remarks', '')
                
                # Create or update attendance
                obj, created = StaffAttendance.objects.update_or_create(
                    teacher_id=teacher_id,
                    date=date,
                    defaults={
                        'status': attendance_status,
                        'remarks': remarks,
                        'marked_by': request.user
                    }
                )
                
                if created:
                    created_count += 1
                else:
                    updated_count += 1
        
        return Response({
            'message': 'Attendance marked successfully',
            'date': date,
            'created': created_count,
            'updated': updated_count,
            'total': created_count + updated_count
        })
    
    @action(detail=False, methods=['get'])
    def daily_summary(self, request):
        """
        Get attendance summary for a specific date
        """
        date = request.query_params.get('date', timezone.now().date())
        
        total_staff = Teacher.objects.filter(is_active=True).count()
        
        attendance = StaffAttendance.objects.filter(date=date)
        
        present = attendance.filter(status='present').count()
        absent = attendance.filter(status='absent').count()
        late = attendance.filter(status='late').count()
        on_leave = attendance.filter(status='on_leave').count()
        
        not_marked = total_staff - attendance.count()
        
        return Response({
            'date': date,
            'total_staff': total_staff,
            'present': present,
            'absent': absent,
            'late': late,
            'on_leave': on_leave,
            'not_marked': not_marked,
            'attendance_rate': round((present / total_staff * 100), 2) if total_staff > 0 else 0
        })
    
    @action(detail=False, methods=['get'])
    def monthly_summary(self, request):
        """
        Get monthly attendance summary for all staff
        """
        month = request.query_params.get('month', timezone.now().month)
        year = request.query_params.get('year', timezone.now().year)
        
        month = int(month)
        year = int(year)
        
        # Get all active teachers
        teachers = Teacher.objects.filter(is_active=True)
        
        summary_data = []
        
        for teacher in teachers:
            attendance = StaffAttendance.objects.filter(
                teacher=teacher,
                date__month=month,
                date__year=year
            )
            
            present = attendance.filter(status='present').count()
            absent = attendance.filter(status='absent').count()
            late = attendance.filter(status='late').count()
            on_leave = attendance.filter(status='on_leave').count()
            
            total_marked = attendance.count()
            
            summary_data.append({
                'teacher_id': teacher.id,
                'employee_id': teacher.employee_id if hasattr(teacher, 'employee_id') else '',
                'name': teacher.user.get_full_name(),
                'present': present,
                'absent': absent,
                'late': late,
                'on_leave': on_leave,
                'total_marked': total_marked,
                'attendance_percentage': round((present / total_marked * 100), 2) if total_marked > 0 else 0
            })
        
        return Response({
            'month': month,
            'year': year,
            'staff_count': teachers.count(),
            'summary': summary_data
        })
    
    @action(detail=False, methods=['get'])
    def export_monthly_report(self, request):
        """
        Export monthly attendance report to Excel
        """
        month = request.query_params.get('month', timezone.now().month)
        year = request.query_params.get('year', timezone.now().year)
        
        month = int(month)
        year = int(year)
        
        teachers = Teacher.objects.filter(is_active=True)
        
        # Create workbook
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f'Attendance {month}-{year}'
        
        # Header
        headers = [
            'Employee ID', 'Name', 'Present', 'Absent', 'Late',
            'On Leave', 'Total Marked', 'Attendance %'
        ]
        
        for col_num, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col_num, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color='366092', end_color='366092', fill_type='solid')
            cell.font = Font(bold=True, color='FFFFFF')
        
        # Data rows
        for row_num, teacher in enumerate(teachers, start=2):
            attendance = StaffAttendance.objects.filter(
                teacher=teacher,
                date__month=month,
                date__year=year
            )
            
            present = attendance.filter(status='present').count()
            absent = attendance.filter(status='absent').count()
            late = attendance.filter(status='late').count()
            on_leave = attendance.filter(status='on_leave').count()
            total_marked = attendance.count()
            
            attendance_pct = round((present / total_marked * 100), 2) if total_marked > 0 else 0
            
            ws.cell(row=row_num, column=1, value=teacher.employee_id if hasattr(teacher, 'employee_id') else '')
            ws.cell(row=row_num, column=2, value=teacher.user.get_full_name())
            ws.cell(row=row_num, column=3, value=present)
            ws.cell(row=row_num, column=4, value=absent)
            ws.cell(row=row_num, column=5, value=late)
            ws.cell(row=row_num, column=6, value=on_leave)
            ws.cell(row=row_num, column=7, value=total_marked)
            ws.cell(row=row_num, column=8, value=attendance_pct)
        
        # Save
        output = BytesIO()
        wb.save(output)
        output.seek(0)
        
        response = HttpResponse(
            output.read(),
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        )
        response['Content-Disposition'] = f'attachment; filename="staff_attendance_{month}_{year}.xlsx"'
        
        return response


# ==================== DOUBLE-ENTRY ACCOUNTING ====================

class AccountingSystemEnhancedViewSet(viewsets.ViewSet):
    """
    Complete double-entry accounting system
    Chart of Accounts → Ledgers → Vouchers → Trial Balance → Reports
    """
    permission_classes = [IsAuthenticated]
    
    @action(detail=False, methods=['get'])
    def chart_of_accounts(self, request):
        """
        Get complete chart of accounts organized by type
        """
        accounts = ChartOfAccount.objects.all().order_by('account_type', 'code')
        
        organized = {}
        for account in accounts:
            acc_type = account.get_account_type_display()
            if acc_type not in organized:
                organized[acc_type] = []
            
            organized[acc_type].append({
                'id': account.id,
                'code': account.code,
                'name': account.name,
                'balance': float(account.balance)
            })
        
        return Response({
            'chart_of_accounts': organized,
            'total_accounts': accounts.count()
        })
    
    @action(detail=False, methods=['post'])
    def create_journal_entry(self, request):
        """
        Create a journal entry with double-entry validation
        """
        entry_date = request.data.get('entry_date')
        description = request.data.get('description')
        lines = request.data.get('lines', [])
        
        if not entry_date or not description or not lines:
            return Response(
                {'error': 'entry_date, description, and lines are required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        # Validate double-entry
        total_debit = Decimal('0.00')
        total_credit = Decimal('0.00')
        
        for line in lines:
            total_debit += Decimal(str(line.get('debit', 0)))
            total_credit += Decimal(str(line.get('credit', 0)))
        
        if abs(float(total_debit) - float(total_credit)) > 0.01:
            return Response(
                {'error': f'Debits ({total_debit}) must equal Credits ({total_credit})'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            with transaction.atomic():
                # Create journal entry
                entry = JournalEntry.objects.create(
                    entry_date=entry_date,
                    description=description,
                    total_debit=total_debit,
                    total_credit=total_credit,
                    created_by=request.user
                )
                
                # Create journal entry lines
                for line in lines:
                    JournalEntryLine.objects.create(
                        journal_entry=entry,
                        account_id=line.get('account_id'),
                        description=line.get('description', ''),
                        debit=Decimal(str(line.get('debit', 0))),
                        credit=Decimal(str(line.get('credit', 0)))
                    )
                
                return Response({
                    'message': 'Journal entry created successfully',
                    'entry_number': entry.entry_number,
                    'total_debit': float(total_debit),
                    'total_credit': float(total_credit)
                })
        
        except Exception as e:
            return Response(
                {'error': f'Failed to create journal entry: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['post'])
    def post_journal_entry(self, request):
        """
        Post a journal entry and update account balances
        """
        entry_id = request.data.get('entry_id')
        
        if not entry_id:
            return Response(
                {'error': 'entry_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            entry = JournalEntry.objects.get(id=entry_id)
        except JournalEntry.DoesNotExist:
            return Response(
                {'error': 'Journal entry not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        if entry.status == 'posted':
            return Response(
                {'error': 'Journal entry already posted'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            with transaction.atomic():
                # Update account balances
                for line in entry.lines.all():
                    account = line.account
                    
                    # Debit increases assets/expenses, decreases liabilities/income/equity
                    # Credit increases liabilities/income/equity, decreases assets/expenses
                    
                    if account.account_type in ['asset', 'expense']:
                        account.balance += line.debit - line.credit
                    else:  # liability, income, equity
                        account.balance += line.credit - line.debit
                    
                    account.save()
                
                # Mark entry as posted
                entry.status = 'posted'
                entry.posted_at = timezone.now()
                entry.save()
                
                return Response({
                    'message': 'Journal entry posted successfully',
                    'entry_number': entry.entry_number
                })
        
        except Exception as e:
            return Response(
                {'error': f'Failed to post journal entry: {str(e)}'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR
            )
    
    @action(detail=False, methods=['get'])
    def ledger(self, request):
        """
        Get ledger for a specific account
        """
        account_id = request.query_params.get('account_id')
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date')
        
        if not account_id:
            return Response(
                {'error': 'account_id is required'},
                status=status.HTTP_400_BAD_REQUEST
            )
        
        try:
            account = ChartOfAccount.objects.get(id=account_id)
        except ChartOfAccount.DoesNotExist:
            return Response(
                {'error': 'Account not found'},
                status=status.HTTP_404_NOT_FOUND
            )
        
        # Get journal entry lines for this account
        lines = JournalEntryLine.objects.filter(
            account=account,
            journal_entry__status='posted'
        ).select_related('journal_entry')
        
        if start_date:
            lines = lines.filter(journal_entry__entry_date__gte=start_date)
        if end_date:
            lines = lines.filter(journal_entry__entry_date__lte=end_date)
        
        lines = lines.order_by('journal_entry__entry_date')
        
        ledger_entries = []
        running_balance = Decimal('0.00')
        
        for line in lines:
            running_balance += line.debit - line.credit
            
            ledger_entries.append({
                'date': line.journal_entry.entry_date,
                'entry_number': line.journal_entry.entry_number,
                'description': line.description or line.journal_entry.description,
                'debit': float(line.debit),
                'credit': float(line.credit),
                'balance': float(running_balance)
            })
        
        return Response({
            'account_code': account.code,
            'account_name': account.name,
            'account_type': account.get_account_type_display(),
            'current_balance': float(account.balance),
            'ledger_entries': ledger_entries
        })
    
    @action(detail=False, methods=['get'])
    def trial_balance(self, request):
        """
        Generate trial balance report
        """
        as_of_date = request.query_params.get('as_of_date', timezone.now().date())
        
        accounts = ChartOfAccount.objects.all().order_by('account_type', 'code')
        
        trial_balance_data = []
        total_debit = Decimal('0.00')
        total_credit = Decimal('0.00')
        
        for account in accounts:
            balance = account.balance
            
            # Determine debit or credit side
            if account.account_type in ['asset', 'expense']:
                # Normal debit balance
                if balance >= 0:
                    debit = balance
                    credit = Decimal('0.00')
                else:
                    debit = Decimal('0.00')
                    credit = abs(balance)
            else:  # liability, income, equity
                # Normal credit balance
                if balance >= 0:
                    debit = Decimal('0.00')
                    credit = balance
                else:
                    debit = abs(balance)
                    credit = Decimal('0.00')
            
            total_debit += debit
            total_credit += credit
            
            trial_balance_data.append({
                'code': account.code,
                'name': account.name,
                'type': account.get_account_type_display(),
                'debit': float(debit),
                'credit': float(credit)
            })
        
        is_balanced = abs(float(total_debit) - float(total_credit)) < 0.01
        
        return Response({
            'as_of_date': as_of_date,
            'trial_balance': trial_balance_data,
            'total_debit': float(total_debit),
            'total_credit': float(total_credit),
            'is_balanced': is_balanced,
            'difference': float(total_debit - total_credit)
        })
    
    @action(detail=False, methods=['get'])
    def profit_loss_statement(self, request):
        """
        Generate Profit & Loss (Income Statement) report
        """
        start_date = request.query_params.get('start_date')
        end_date = request.query_params.get('end_date', timezone.now().date())
        
        # Get income accounts
        income_accounts = ChartOfAccount.objects.filter(account_type='income')
        total_income = income_accounts.aggregate(Sum('balance'))['balance__sum'] or Decimal('0.00')
        
        # Get expense accounts
        expense_accounts = ChartOfAccount.objects.filter(account_type='expense')
        total_expenses = expense_accounts.aggregate(Sum('balance'))['balance__sum'] or Decimal('0.00')
        
        # Calculate profit/loss
        net_profit = total_income - total_expenses
        
        return Response({
            'period': f'{start_date} to {end_date}',
            'income': {
                'accounts': [{'code': acc.code, 'name': acc.name, 'amount': float(acc.balance)} for acc in income_accounts],
                'total': float(total_income)
            },
            'expenses': {
                'accounts': [{'code': acc.code, 'name': acc.name, 'amount': float(acc.balance)} for acc in expense_accounts],
                'total': float(total_expenses)
            },
            'net_profit_loss': float(net_profit),
            'is_profit': net_profit >= 0
        })
    
    @action(detail=False, methods=['get'])
    def balance_sheet(self, request):
        """
        Generate Balance Sheet report
        """
        as_of_date = request.query_params.get('as_of_date', timezone.now().date())
        
        # Assets
        assets = ChartOfAccount.objects.filter(account_type='asset')
        total_assets = assets.aggregate(Sum('balance'))['balance__sum'] or Decimal('0.00')
        
        # Liabilities
        liabilities = ChartOfAccount.objects.filter(account_type='liability')
        total_liabilities = liabilities.aggregate(Sum('balance'))['balance__sum'] or Decimal('0.00')
        
        # Equity
        equity = ChartOfAccount.objects.filter(account_type='equity')
        total_equity = equity.aggregate(Sum('balance'))['balance__sum'] or Decimal('0.00')
        
        # Total liabilities + equity
        total_liabilities_equity = total_liabilities + total_equity
        
        return Response({
            'as_of_date': as_of_date,
            'assets': {
                'accounts': [{'code': acc.code, 'name': acc.name, 'amount': float(acc.balance)} for acc in assets],
                'total': float(total_assets)
            },
            'liabilities': {
                'accounts': [{'code': acc.code, 'name': acc.name, 'amount': float(acc.balance)} for acc in liabilities],
                'total': float(total_liabilities)
            },
            'equity': {
                'accounts': [{'code': acc.code, 'name': acc.name, 'amount': float(acc.balance)} for acc in equity],
                'total': float(total_equity)
            },
            'total_liabilities_equity': float(total_liabilities_equity),
            'is_balanced': abs(float(total_assets) - float(total_liabilities_equity)) < 0.01
        })


# ==================== Continue in next file due to length ====================
# Part 2 will include: Fee Management, Wallet System, Inventory, Library, Transport, Dormitory
